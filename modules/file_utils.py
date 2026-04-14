"""
modules/file_utils.py
Excel / CSV helpers: sheet names, dimensions, merged-cell metadata, totals rows.
"""

import csv
import os
import re

import openpyxl
from openpyxl.utils import get_column_letter

from modules.cell_format import format_cell_value_with_fmt


# ── Sheet enumeration ─────────────────────────────────────────────────────────

def get_sheet_names(file_path: str) -> list[str]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return ["Sheet1"]
    wb      = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    names   = list(wb.sheetnames)
    wb.close()
    summary = [n for n in names if n.strip().lower() == "summary"]
    others  = [n for n in names if n.strip().lower() != "summary"]
    return summary + others


# ── Dimensions ────────────────────────────────────────────────────────────────

def get_sheet_dimensions(file_path: str, sheet_name: str) -> tuple[int, int]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        return len(rows), max((len(r) for r in rows), default=0)
    wb    = openpyxl.load_workbook(file_path, data_only=True)
    ws    = wb[sheet_name]
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    if max_r == 0 or max_c == 0:
        actual_rows = actual_cols = 0
        for row in ws.iter_rows():
            if any(cell.value is not None for cell in row):
                actual_rows += 1
                row_col = max(
                    (cell.column for cell in row if cell.value is not None), default=0
                )
                actual_cols = max(actual_cols, row_col)
        max_r, max_c = actual_rows, actual_cols
    wb.close()
    return max_r, max_c


# ── Merged-cell metadata ──────────────────────────────────────────────────────

def extract_merged_cell_metadata(file_path: str, sheet_name: str) -> dict:
    """
    Extract every merged region from the sheet.
    The full cell value is stored without any truncation — display truncation
    is the UI's responsibility, not this function's.
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return {}
    wb          = openpyxl.load_workbook(file_path, data_only=True)
    ws          = wb[sheet_name]
    merged_info = {}
    for mr in ws.merged_cells.ranges:
        mn_r, mn_c, mx_r, mx_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        cell       = ws.cell(mn_r, mn_c)
        # Store the complete, untruncated value — [:35] cuts happen only in UI widgets
        val        = str(cell.value).strip() if cell.value else ""
        span_cols  = mx_c - mn_c + 1
        span_rows  = mx_r - mn_r + 1
        region_type = (
            "TITLE"  if mn_r <= 3 and span_cols >= 3 else
            "HEADER" if span_cols >= 2 and span_rows == 1 else
            "DATA"
        )
        merged_info[f"R{mn_r}C{mn_c}"] = {
            "value": val, "type": region_type,
            "row_start": mn_r, "col_start": mn_c,
            "row_end": mx_r,   "col_end":   mx_c,
            "span_cols": span_cols, "span_rows": span_rows,
            "excel_row": mn_r,      "excel_col": mn_c,
        }
    wb.close()
    return merged_info


# ── Financial field detection ─────────────────────────────────────────────────

# Words that confirm a field holds a monetary / numeric amount
_STRONG_AMOUNT_WORDS = re.compile(
    r"\b(paid|incurred|reserve|outstanding|cost|costs|expense|"
    r"recovery|recoveries|subrogation|deductible|fee|indemnity|"
    r"damage|settlement|award|judgment|medical|legal|defense|adjusting|"
    r"amount|subtotal|sum|balance)\b",
    re.IGNORECASE,
)

_FINANCIAL_KEYWORDS = re.compile(
    r"\b(paid|incurred|loss|reserve|outstanding|cost|costs|expense|"
    r"recovery|recoveries|subrogation|deductible|tpa|fee|indemnity|"
    r"damage|settlement|award|judgment|medical|legal|defense|adjusting|"
    r"amount|total|subtotal|sum|balance)\b",
    re.IGNORECASE,
)

# Words that mark a column as an identifier / category, not an amount
_IDENTIFIER_KEYWORDS = re.compile(
    r"\b(num|number|id|no\.?|ref|code|pol|policy|file|claim\s*id|"
    r"name|assured|claimant|insured|adjuster|employer|"
    r"dol|status|home|state|jurisdiction|address|cause|"
    r"description|narrative|period|contract|phone|email|zip|"
    r"county|country|city|unit|type|category|class|lob|line)\b",
    re.IGNORECASE,
)

# Leading tokens that signal a field is financial even when identifier words follow
# (e.g. "LOSS PAID TO DATE" starts with "loss"/"paid" so it's financial despite "date")
_LEADING_FINANCIAL_TOKENS = {
    "loss", "paid", "total", "outstanding", "adjusting",
    "tpa", "defense", "expense", "recovery", "recoveries",
    "reserve", "indemnity", "medical", "legal", "cost",
    "costs", "current", "gross", "net",
}


def is_financial_field(field_name: str) -> bool:
    """
    Return True if *field_name* represents a financial amount column that
    should be included in sheet totals.

    Design principles — nothing is hardcoded to specific column names:
    - Works entirely on keyword semantics derived from the column label.
    - Handles raw legacy names ("LOSS PAID TO DATE"), standardised names
      ("Total Incurred"), and duplicate-suffixed names ("ADJUSTING COSTS_2").
    - Date fields are always excluded even when they contain financial words
      (e.g. "Loss Date", "Date of Loss") unless a strong amount word also
      appears (e.g. "LOSS PAID TO DATE" → financial).
    - Identifier fields (claim numbers, policy numbers, names, statuses)
      are always excluded even though they may hold numeric-looking values.
    """
    # Strip trailing _2, _3 … suffixes produced by duplicate-column disambiguation
    f = re.sub(r'_\d+$', '', field_name.strip())

    # Date/timestamp fields → never a summed financial amount
    # Exception: if a strong amount word also appears (e.g. "LOSS PAID TO DATE")
    if re.search(r'\b(date|dt)\b', f, re.IGNORECASE):
        if not _STRONG_AMOUNT_WORDS.search(f):
            return False

    # Identifier/category fields → excluded unless the first meaningful token
    # is a known financial lead word (handles "LOSS PAID TO DATE" which starts
    # with "loss" but also contains "date" handled above)
    if _IDENTIFIER_KEYWORDS.search(f):
        tokens = re.split(r'[\s_]+', f.lower())
        first_meaningful = next((t for t in tokens if len(t) > 1), "")
        if first_meaningful not in _LEADING_FINANCIAL_TOKENS:
            return False

    return bool(_FINANCIAL_KEYWORDS.search(f))


# ── Totals-row extraction ─────────────────────────────────────────────────────

def extract_totals_row(file_path: str, sheet_name: str) -> dict:
    ext    = os.path.splitext(file_path)[1].lower()
    totals = {}
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        cell_rows = None
    else:
        wb        = openpyxl.load_workbook(file_path, data_only=True)
        ws        = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        rows      = raw_rows
        wb.close()
    if not rows:
        return totals

    header_row_index, headers = None, []
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if "claim" in row_text and ("date" in row_text or "incurred" in row_text or "paid" in row_text):
            header_row_index = i
            headers = [str(h).strip() if h is not None else f"Column_{j}" for j, h in enumerate(row)]
            break
    if header_row_index is None or not headers:
        return totals

    totals_rows = []
    for r_idx_rel, raw_row in enumerate(rows[header_row_index + 1:]):
        r_idx    = header_row_index + 2 + r_idx_rel
        if not any(raw_row):
            continue
        row_text = " ".join([str(c).lower() for c in raw_row if c])
        if any(kw in row_text for kw in ["total", "subtotal", "grand total", "sum", "totals"]):
            row_data: dict = {}
            cell_row = cell_rows[header_row_index + 1 + r_idx_rel] if cell_rows else None
            for c_idx_0, raw_val in enumerate(raw_row):
                if c_idx_0 >= len(headers):
                    continue
                if cell_row and c_idx_0 < len(cell_row):
                    clean_val = format_cell_value_with_fmt(cell_row[c_idx_0])
                    real_col  = cell_row[c_idx_0].column if hasattr(cell_row[c_idx_0], "column") else c_idx_0 + 1
                else:
                    clean_val = str(raw_val).strip() if raw_val is not None else ""
                    real_col  = c_idx_0 + 1
                if clean_val:
                    row_data[headers[c_idx_0]] = {
                        "value": clean_val, "excel_row": r_idx, "excel_col": real_col,
                    }
            if row_data:
                totals_rows.append(row_data)

    if totals_rows:
        totals["rows"]      = totals_rows
        totals["excel_row"] = totals_rows[0].get(list(totals_rows[0].keys())[0], {}).get("excel_row", 9999)
        agg: dict = {}
        for row_data in totals_rows:
            for field, info in row_data.items():
                # Only aggregate genuine financial amount fields
                if not is_financial_field(field):
                    continue
                try:
                    num = float(str(info["value"]).replace(",", "").replace("$", ""))
                    agg[field] = agg.get(field, 0.0) + num
                except Exception:
                    pass
        totals["aggregated"] = {k: round(v, 2) for k, v in agg.items()}
    return totals


# ── Compute totals from parsed claim data (always available) ──────────────────

def compute_totals_from_claims(claims_data: list[dict]) -> dict:
    """
    Calculate column totals directly from the parsed claim rows.

    Only genuine financial amount columns are summed — identifiers like claim
    numbers, policy numbers, names, dates and status fields are excluded using
    semantic keyword analysis (is_financial_field). No column positions or
    names are hardcoded; the filter works on whatever column labels the source
    file uses.

    Returns:
        {
            "aggregated": { field: rounded_float, ... },
            "source":     "computed",
        }
    Always returns at least {"aggregated": {}, "source": "computed"}.
    """
    if not claims_data:
        return {"aggregated": {}, "source": "computed"}

    agg: dict[str, float] = {}

    for claim in claims_data:
        for field, info in claim.items():
            # Skip non-financial fields (identifiers, dates, categories, etc.)
            if not is_financial_field(field):
                continue

            raw = info.get("modified") or info.get("value", "")
            if raw is None:
                continue
            cleaned = str(raw).strip().replace(",", "").replace("$", "").replace("%", "")
            try:
                num = float(cleaned)
            except (ValueError, TypeError):
                continue

            agg[field] = round(agg.get(field, 0.0) + num, 2)

    # Drop columns that summed to exactly zero AND have zero on every row —
    # these are columns the source file included but which carry no data.
    # We keep zero-valued columns that exist because some claims DO have values
    # (e.g. "LOSS PAID THIS MONTH" is legitimately all-zero if nothing was paid
    # this month — still meaningful to show). We only suppress if EVERY value
    # was zero across ALL claims.
    all_zero_fields = set()
    for field in list(agg.keys()):
        if agg[field] == 0.0:
            # Check if this field has any non-zero value across claims
            has_nonzero = any(
                _safe_float(c.get(field, {}).get("modified") or c.get(field, {}).get("value", "")) != 0.0
                for c in claims_data
                if field in c
            )
            if not has_nonzero:
                all_zero_fields.add(field)

    filtered_agg = {k: v for k, v in agg.items() if k not in all_zero_fields}

    return {"aggregated": filtered_agg, "source": "computed"}


def _safe_float(val) -> float:
    """Convert a value to float safely, returning 0.0 on failure."""
    if val is None:
        return 0.0
    try:
        return float(str(val).strip().replace(",", "").replace("$", "").replace("%", ""))
    except (ValueError, TypeError):
        return 0.0


def get_totals_for_sheet(
    file_path: str,
    sheet_name: str,
    claims_data: list[dict],
) -> dict:
    """
    Always return a populated totals dict for the sheet.

    Strategy:
      1. Try to find an explicit totals row in the Excel file.
      2. If none found (or aggregated is empty), compute from parsed claim rows.

    The returned dict always contains an "aggregated" key.
    """
    totals = extract_totals_row(file_path, sheet_name)

    if totals.get("aggregated"):
        totals.setdefault("source", "excel_row")
        return totals

    return compute_totals_from_claims(claims_data)