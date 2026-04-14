"""
modules/parsing.py
Excel / CSV ingestion: classify sheet type, parse rows into list-of-dicts,
skip aggregate/totals rows.

Handles both clean structured layouts and messy legacy "print-style" formats
where:
  - Column headers span two rows (group label + sub-label)
  - Each claim has a sub-row with the address (col B) and cause-of-loss (col D)
  - Section separator rows contain only '----------' dashes
  - Section subtotal rows start with 'Total ...'

NEW (title extraction):
  extract_sheet_title_kvs() scans the pre-header rows of any sheet and returns
  a dict of canonical key → value pairs (TPA Name, Treaty, Cedant,
  Valuation Date, Sheet Title, Sheet Name …).  extract_from_excel() now
  returns this dict as a third element so callers can pass it downstream to
  schema_mapping.extract_title_fields_from_kvs().
"""

import csv
import os
import re

import openpyxl

from modules.cell_format import format_cell_value_with_fmt


# ── Sheet classifier ──────────────────────────────────────────────────────────

"""
def classify_sheet(rows) -> str:
    text = " ".join(str(cell).lower() for row in rows[:20] for cell in row if cell)
    if "line of business" in text:
        return "SUMMARY"
    has_claim = any(x in text for x in [
        "claim number", "claim no", "claim #", "claim id", "claim_id",
        "claim ref", "claimant", "file number", "file no", "file num",
    ])
    has_loss = any(x in text for x in [
        "loss date", "date of loss", "loss dt", "accident date",
        "occurrence date", "incident date", "date of injury", "date of incident",
        "injury date", "dol",
    ])
    has_fin = any(x in text for x in [
        "incurred", "paid", "reserve", "outstanding",
        "total paid", "total incurred", "indemnity", "expense",
    ])
    if has_claim and (has_loss or has_fin):
        return "LOSS_RUN"
    if "policy" in text and ("claim" in text or "incurred" in text):
        return "COMMERCIAL_LOSS_RUN"
    if has_claim:
        return "LOSS_RUN"
    return "UNKNOWN
    """
   

def classify_sheet(rows) -> str:
    text = " ".join(str(cell).lower() for row in rows[:20] for cell in row if cell)

    # SUMMARY detection: "line of business" must co-occur with summary-specific
    # signals, not just appear as a data column header in a loss-run sheet.
    if "line of business" in text:
        summary_co_signals = [
            "# claims", "num claims", "number of claims", "claim count",
            "loss ratio", "loss rate", "frequency", "severity",
        ]
        if any(sig in text for sig in summary_co_signals):
            return "SUMMARY"
        # Also SUMMARY if "line of business" is a standalone first-cell row label
        for row in rows[:20]:
            non_empty = [v for v in row if v is not None and str(v).strip()]
            if non_empty and str(non_empty[0]).lower().strip() == "line of business" and len(non_empty) == 1:
                return "SUMMARY"

    has_claim = any(x in text for x in [
        "claim number", "claim no", "claim #", "claim id", "claim_id",
        "claim ref", "claimant", "file number", "file no", "file num",
        "file ref",
    ])
    has_loss = any(x in text for x in [
        "loss date", "date of loss", "loss dt", "accident date",
        "occurrence date", "incident date", "date of injury", "date of incident",
        "injury date", "dol",
    ])
    has_fin = any(x in text for x in [
        "incurred", "paid", "reserve", "outstanding",
        "total paid", "total incurred", "indemnity", "expense",
    ])
    if has_claim and (has_loss or has_fin):
        return "LOSS_RUN"
    if "policy" in text and ("claim" in text or "incurred" in text):
        return "COMMERCIAL_LOSS_RUN"
    if has_claim:
        return "LOSS_RUN"
    return "UNKNOWN"

# ── Legacy-layout detector ────────────────────────────────────────────────────

def _is_legacy_print_layout(rows: list) -> bool:
    """
    Detect a legacy print-style TPA loss run layout.
    Signatures:
      - Two adjacent rows that together form the column header (neither row
        alone passes _find_header_row's threshold, but together they do).
      - Rows consisting entirely of '----------' dashes appear in the data area.
      - Sub-rows where only columns B and/or D have values (address + cause).
    """
    # Check for '----------' separator rows
    for row in rows:
        non_empty = [c for c in row if c is not None]
        if non_empty and all(str(c).strip() == "----------" for c in non_empty):
            return True

    # Check for the characteristic 2-row header pattern:
    # row N has group labels in some cols, row N+1 has sub-labels in ALL cols
    for i in range(min(20, len(rows) - 1)):
        r1_vals = [str(c).strip() for c in rows[i] if c is not None]
        r2_vals = [str(c).strip() for c in rows[i + 1] if c is not None]
        if len(r2_vals) >= 5 and len(r1_vals) >= 2:
            # r1 is sparse (group labels), r2 is dense (sub-labels)
            r1_filled = sum(1 for c in rows[i] if c)
            r2_filled = sum(1 for c in rows[i + 1] if c)
            if r2_filled > r1_filled * 1.5 and r1_filled >= 2:
                combined = " ".join(r1_vals + r2_vals).lower()
                if ("file" in combined or "claim" in combined) and (
                    "paid" in combined or "incurred" in combined or "outstanding" in combined
                ):
                    return True
    return False


def _find_legacy_header_rows(rows: list) -> tuple[int, int] | None:
    """
    For legacy layouts find the two consecutive rows that form the header.
    Returns (row_index_of_group_label_row, row_index_of_sub_label_row) or None.
    """
    for i in range(min(25, len(rows) - 1)):
        r1 = rows[i]
        r2 = rows[i + 1]
        r1_filled = sum(1 for c in r1 if c)
        r2_filled = sum(1 for c in r2 if c)
        if r2_filled < 4:
            continue
        combined = " ".join(
            str(c).lower() for c in list(r1) + list(r2) if c
        )
        if ("file" in combined or "claim" in combined or "assured" in combined) and (
            "paid" in combined or "outstanding" in combined or "incurred" in combined
        ):
            if r1_filled >= 2:
                return (i, i + 1)
            # r1 might be empty (single-row header variant)
            if r2_filled >= 5:
                return (i + 1, i + 1)
    return None


def _merge_two_header_rows(row1: list, row2: list) -> list[str]:
    """
    Combine a group-label row and a sub-label row into one list of column names.
    Duplicate merged names get a numeric suffix (_2, _3 …).
    """
    headers: list[str] = []
    seen: dict[str, int] = {}
    for g, s in zip(row1, row2):
        g_s = str(g).strip() if g else ""
        s_s = str(s).strip() if s else ""
        if g_s and s_s and g_s.upper() != s_s.upper():
            name = f"{g_s} {s_s}"
        elif s_s:
            name = s_s
        elif g_s:
            name = g_s
        else:
            name = ""
        if name:
            seen[name] = seen.get(name, 0) + 1
            if seen[name] > 1:
                name = f"{name}_{seen[name]}"
        headers.append(name)
    return headers


# ── Sub-row / separator / subtotal detectors ─────────────────────────────────

def _is_separator_row(row_values: list) -> bool:
    """Row containing only '----------' dashes (and empty cells)."""
    non_empty = [c for c in row_values if c is not None and str(c).strip()]
    if not non_empty:
        return False
    return all(str(c).strip() == "----------" for c in non_empty)


def _is_subtotal_row(row_values: list) -> bool:
    """Row whose first non-empty cell starts with 'Total' (section subtotal)."""
    for c in row_values:
        if c is not None and str(c).strip():
            return bool(re.match(r"^total\b", str(c).strip(), re.IGNORECASE))
    return False


def _is_legacy_sub_row(row_values: list, num_cols: int) -> bool:
    """
    In legacy print layouts each claim is followed by a sub-row containing
    only the address (col B / index 1) and/or cause-of-loss (col D / index 3).
    Signature: col 0 (FILE NUM) is empty, ≤ 3 non-empty cells total, and at
    least one of col 1 or col 3 has a value.
    """
    if not row_values or row_values[0] is not None:
        return False
    non_empty = [c for c in row_values if c is not None and str(c).strip()]
    if len(non_empty) == 0 or len(non_empty) > 3:
        return False
    has_addr_or_cause = (
        (len(row_values) > 1 and row_values[1] is not None) or
        (len(row_values) > 3 and row_values[3] is not None)
    )
    return has_addr_or_cause


# ── Aggregate-row detection ───────────────────────────────────────────────────

_AGGREGATE_PATTERNS = re.compile(
    r"^(total|totals|grand\s*total|subtotal|aggregate|summary|sum|report\s*(date|total|summary)|"
    r"all\s+adjusters|ytd\s+total|period\s+total|fiscal\s+total|portfolio\s+total|"
    r"TOTALS_AGGREGATE|SUMMARY_FLIBBER|AGGREGATE_ZORP|SUMMARY_ZORP)",
    re.IGNORECASE,
)
_AGGREGATE_EXTRA = re.compile(
    r"(aggregate|zorp|flibber|summary|zoop|gorp|totals?_|_total|report_date|all_adjuster)",
    re.IGNORECASE,
)


def _is_aggregate_row(row_values: list) -> bool:
    non_empty = [str(v).strip() for v in row_values if v is not None and str(v).strip()]
    if not non_empty:
        return False
    first_val = non_empty[0]
    if _AGGREGATE_PATTERNS.match(first_val):
        return True
    if _AGGREGATE_EXTRA.search(first_val):
        return True
    first_tokens    = re.split(r"[_\s]+", first_val.lower())
    aggregate_tokens = {"total", "totals", "aggregate", "summary", "subtotal", "grand", "portfolio", "report"}
    if len(first_tokens) >= 2 and any(t in aggregate_tokens for t in first_tokens):
        return True
    for v in non_empty[:6]:
        if re.match(
            r"(total\s+claims|report\s+date|all\s+adjusters|open:\s*\d|pend:\s*\d|open:\d)",
            str(v), re.IGNORECASE,
        ):
            return True
    nums = [float(v) for v in row_values if isinstance(v, (int, float))]
    if nums and len(nums) >= 3 and all(n > 50_000 for n in nums):
        # Exempt rows whose first value looks like a claim/file number:
        # either alphanumeric pattern (e.g. 'AB-1234') or a pure integer id.
        is_claim_id = (
            re.match(r"^[A-Z]{2,5}[-_][A-Z]{0,3}\d{3,}", first_val, re.IGNORECASE)
            or re.match(r"^\d{4,}$", first_val.strip())
        )
        if not is_claim_id:
            return True
    return False


# ── Sheet title / metadata extractor ─────────────────────────────────────────

# Semantic label map: normalise raw label text → canonical key name
_LABEL_ALIASES: dict[str, str] = {
    # TPA / reinsurer
    "prepared for":     "Reinsurer",
    "reinsurer":        "Reinsurer",
    "prepared by":      "TPA Name",
    # Treaty / program
    "treaty":           "Treaty",
    "program":          "Treaty",
    "policy":           "Policy Number",
    # Cedant
    "cedant":           "Cedant",
    "ceding company":   "Cedant",
    "insurer":          "Cedant",
    # Dates
    "valuation date":   "Valuation Date",
    "valuation":        "Valuation Date",
    "as of":            "Valuation Date",
    "report date":      "Report Date",
    "report generated": "Report Date",
    "effective date":   "Effective Date",
    # Identifiers
    "policy number":    "Policy Number",
    "policy no":        "Policy Number",
    "policy #":         "Policy Number",
    "insured":          "Insured Name",
    "named insured":    "Insured Name",
    # Coverage / LOB
    "line of business": "Line of Business",
    "lob":              "Line of Business",
    "coverage":         "Coverage Type",
}


def _canonical_label(raw: str) -> str | None:
    """Map a raw label string to a canonical field name, or None if unrecognised."""
    key = raw.strip().rstrip(":").lower()
    return _LABEL_ALIASES.get(key)


def _try_inline_kv(cell_text: str) -> list[tuple[str, str]]:
    """
    Parse a single cell that contains one or more 'Key: Value' fragments.
    E.g. "Treaty: Casualty Surplus Lines 2025" or "Cedant: Hartford Financial".
    Returns a list of (raw_label, value) pairs.
    """
    pairs = []
    # Split on long whitespace gaps or pipe characters that separate pairs on same cell
    segments = re.split(r'\s{3,}|\|', str(cell_text))
    for seg in segments:
        m = re.match(r'^([A-Za-z][^:]{0,40}):\s*(.+)$', seg.strip())
        if m:
            pairs.append((m.group(1).strip(), m.group(2).strip()))
    return pairs


def extract_sheet_title_kvs(
    raw_rows: list,
    cell_rows: list,
    header_row_idx: int | None,
    sheet_name: str,
) -> dict:
    """
    Extract key-value metadata from the pre-header title area of a sheet.

    Scans every row above ``header_row_idx`` (or the first 15 rows when the
    sheet has no recognisable column header) and returns a dict of
    canonical-key → info-dict pairs, e.g.::

        {
            "TPA Name":       {"value": "Heritage Risk Consultants", "excel_row": 1, ...},
            "Sheet Title":    {"value": "Program Year 2025",         "excel_row": 2, ...},
            "Reinsurer":      {"value": "Munich Reinsurance …",      "excel_row": 3, ...},
            "Valuation Date": {"value": "12/31/2025",                "excel_row": 3, ...},
            "Treaty":         {"value": "Property Cat XL 2020-2025", "excel_row": 4, ...},
            "Cedant":         {"value": "Chubb Limited",             "excel_row": 4, ...},
            "Sheet Name":     {"value": "Loss Run 2025",             "excel_row": 0, ...},
        }

    Three cell patterns are handled (all without any hardcoding):

    * **Pattern A** – lone title rows (single non-empty cell, no colon):
      row 0 → ``TPA Name``; subsequent rows → ``Sheet Title`` (with optional
      LOB extraction from "Loss Run Report — <lob>" phrasing).
    * **Pattern B** – inline ``Key: Value`` in a single cell
      (e.g. "Treaty: Casualty Surplus Lines 2025").
    * **Pattern C** – multi-cell label/value pairs on the same row
      (e.g. col 0 = "Prepared For:", col 1 = "Munich Re …",
             col 4 = "Valuation Date:", col 5 = "12/31/2025").

    The sheet tab name is always stored as ``"Sheet Name"`` (source = "sheet_tab").
    """
    # scan_limit = header_row_idx if header_row_idx is not None else min(15, len(raw_rows))
    scan_limit = max(header_row_idx or 0, 15)
    found: dict = {}

    def _store(canonical: str, value: str, excel_row: int, excel_col: int):
        """Store a KV pair only if the canonical key is not yet recorded."""
        if canonical not in found and str(value).strip():
            found[canonical] = {
                "value":     str(value).strip(),
                "original":  str(value).strip(),
                "modified":  str(value).strip(),
                "source":    "title_kv",
                "excel_row": excel_row,
                "excel_col": excel_col,
            }

    for r_idx, row in enumerate(raw_rows[:scan_limit]):
        excel_row = r_idx + 1  # 1-based

        # Collect (col_index, value) for non-empty cells
        non_empty = [
            (c_idx, v) for c_idx, v in enumerate(row)
            if v is not None and str(v).strip()
        ]
        if not non_empty:
            continue

        # ── Pattern A: lone title row (one cell, no colon) ───────────────────
        if len(non_empty) == 1:
            c_idx, val = non_empty[0]
            val_s = str(val).strip()
            # Skip cells that look purely numeric / formula results
            if re.match(r'^[\d$,()\-\.]+$', val_s):
                continue
            if ":" not in val_s:
                if r_idx == 0:
                    # First data row is almost always the TPA / company name.
                    # Strip "— Sheet Subtitle" or "- Report Type" suffixes.
                    tpa_name = re.split(r'\s*[\u2014\u2013]\s*', val_s)[0].strip()
                    if ' - ' in tpa_name:
                        parts = tpa_name.split(' - ', 1)
                        if re.search(r'\b(report|run|detail|summary|schedule|listing)\b',
                                     parts[1], re.IGNORECASE):
                            tpa_name = parts[0].strip()
                    _store("TPA Name", tpa_name, excel_row, c_idx + 1)
                else:
                    # Try to extract LOB from phrasing like
                    # "Loss Run Report — GL - General Liability"
                    # "Annual Loss Run — Program Year 2025"
                    lob_match = re.search(
                        r'(?:loss\s+run\s+report\s*[—\-–]+\s*'
                        r'|annual\s+loss\s+run\s*[—\-–]+\s*'
                        r'|program\s+year\s+\d{4}\s*[—\-–]?\s*)(.+)',
                        val_s, re.IGNORECASE,
                    )
                    if lob_match:
                        _store("Sheet Title", lob_match.group(1).strip(), excel_row, c_idx + 1)
                    else:
                        _store("Sheet Title", val_s, excel_row, c_idx + 1)
                continue  # Pattern A handled; move to next row

        # ── Pattern B: inline "Key: Value" scan — applied to every cell ─────
        # No longer restricted to rows with ≤ 2 cells: any pre-header cell may
        # contain a self-contained "Key: Value" string (e.g. TPA_14 row 2 has
        # "Treaty: Casualty Surplus Lines 2025" and "Cedant: Hartford Financial"
        # in two non-adjacent cells, and TPA_14 row 3 has "Valuation: 03/31/2025"
        # as a single cell). We scan every non-empty cell in the row.
        for c_idx, val in non_empty:
            val_s = str(val).strip()
            if ":" in val_s and not re.match(r'^\d', val_s):
                for raw_label, raw_value in _try_inline_kv(val_s):
                    canonical = _canonical_label(raw_label)
                    if canonical:
                        _store(canonical, raw_value, excel_row, c_idx + 1)

        # ── Pattern C: adjacent label-cell / value-cell pairs ────────────────
        # Scans consecutive (label_cell, value_cell) pairs across all non-empty
        # cells. Skips cells that already contain inline "Key: Value" text
        # (those are fully handled by Pattern B above).
        i = 0
        cells = non_empty
        while i < len(cells) - 1:
            c_label_idx, label_val = cells[i]
            c_value_idx, value_val = cells[i + 1]
            label_s = str(label_val).strip()
            value_s = str(value_val).strip()

            # is_label = (
            #     label_s.endswith(":")
            #     or _canonical_label(label_s) is not None
            # )
            is_label = (
            ":" in label_s
             or label_s.lower().strip().replace(":", "") in _LABEL_ALIASES
            )
            # Skip cells that are themselves inline KV pairs (already handled
            # by Pattern B); they contain ":" but don't end with ":"
            if is_label and ":" in label_s and not label_s.endswith(":"):
                i += 1
                continue

            if is_label:
                canonical = (
                    _canonical_label(label_s.rstrip(":").strip())
                    or _canonical_label(label_s)
                )
             ##------------------   
                if label_s.endswith(":") and (not value_s or not str(value_s).strip()):
                    for j in range(i + 1, len(cells)):
                        next_val = str(cells[j][1]).strip()
                    if next_val:
                        value_s = next_val
                        break
                if canonical:
                    _store(canonical, value_s, excel_row, c_value_idx + 1)
                i += 2
            else:
                i += 1

    # ── Always record the sheet tab name ─────────────────────────────────────
    if "Sheet Name" not in found:
        found["Sheet Name"] = {
            "value":     sheet_name,
            "original":  sheet_name,
            "modified":  sheet_name,
            "source":    "sheet_tab",
            "excel_row": 0,
            "excel_col": 0,
        }

    return found


# ── Main entry point ──────────────────────────────────────────────────────────

def extract_from_excel(
    file_path: str,
    sheet_name: str,
) -> tuple[list, str, dict]:
    """
    Parse a single sheet from an Excel or CSV file.

    Returns
    -------
    (claims, sheet_type, title_kvs)

    ``claims``     – list of row-dicts in the standard ``{field: {value, modified, …}}``
                     format, one entry per claim row.
    ``sheet_type`` – classifier label e.g. "LOSS_RUN", "SUMMARY", "UNKNOWN".
    ``title_kvs``  – dict of canonical metadata extracted from the pre-header
                     title area (TPA Name, Treaty, Cedant, Valuation Date …).
                     Empty dict for CSV files.  Pass this to
                     ``schema_mapping.extract_title_fields_from_kvs()`` to merge
                     it with the existing title-field pipeline.
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        if not rows:
            return [], "UNKNOWN", {}
        claims, sheet_type = parse_rows(classify_sheet(rows), rows)
        return claims, sheet_type, {}
    else:
        wb        = openpyxl.load_workbook(file_path, data_only=True)
        ws        = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        wb.close()
        if not raw_rows:
            return [], "UNKNOWN", {}

        sheet_type = classify_sheet(raw_rows)
        hri        = _find_header_row(raw_rows)
        title_kvs  = extract_sheet_title_kvs(raw_rows, cell_rows, hri, sheet_name)
        claims, sheet_type = parse_rows_with_cells(sheet_type, raw_rows, cell_rows)
        return claims, sheet_type, title_kvs


# ── Row parsers ───────────────────────────────────────────────────────────────

def _find_header_row(rows: list) -> int | None:
    for i, row in enumerate(rows[:20]):
        rt = " ".join([str(c).lower() for c in row if c])
        if (
            "claim" in rt or "employee name" in rt or "driver name" in rt
            or "claimant" in rt or "file" in rt
        ) and (
            "date" in rt or "incurred" in rt or "paid" in rt
            or "injury" in rt or "incident" in rt or "amount" in rt or "reserve" in rt
        ):
            return i
    for i, row in enumerate(rows[:5]):
        if sum(1 for c in row if c) >= 3:
            return i
    return None


def parse_rows_with_cells(sheet_type: str, rows: list, cell_rows: list) -> tuple[list, str]:
    # ── SUMMARY sheet ─────────────────────────────────────────────────────────
    if sheet_type == "SUMMARY":
        hri = None
        for i, row in enumerate(rows[:20]):
            rt = " ".join([str(c).lower() for c in row if c])
            if "sheet" in rt and "line of business" in rt:
                hri = i
                break
        if hri is None:
            return [], sheet_type
        headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        extracted = []
        for r_idx_rel, (raw_row, cell_row) in enumerate(zip(rows[hri + 1:], cell_rows[hri + 1:])):
            r_idx = hri + 2 + r_idx_rel
            if not any(raw_row):
                continue
            row_data: dict = {}
            for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
                if c_idx_0 >= len(headers):
                    continue
                clean_val = format_cell_value_with_fmt(cell)
                real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
                row_data[headers[c_idx_0]] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": real_col,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    # ── Legacy print layout ───────────────────────────────────────────────────
    if _is_legacy_print_layout(rows):
        return _parse_legacy_layout_with_cells(sheet_type, rows, cell_rows)

    # ── Standard LOSS_RUN / COMMERCIAL_LOSS_RUN ───────────────────────────────
    hri = _find_header_row(rows)
    if hri is None:
        return [], sheet_type
    headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
    extracted = []
    for r_idx_rel, (raw_row, cell_row) in enumerate(zip(rows[hri + 1:], cell_rows[hri + 1:])):
        r_idx = hri + 2 + r_idx_rel
        if not any(raw_row):
            continue
        if any(str(c).lower().strip() in ["totals", "total", "grand total", "subtotal"] for c in raw_row if c):
            break
        if _is_aggregate_row(raw_row):
            continue
        row_data: dict = {}
        for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
            if c_idx_0 >= len(headers):
                continue
            clean_val = format_cell_value_with_fmt(cell)
            real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
            row_data[headers[c_idx_0]] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": real_col,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


def _parse_legacy_layout_with_cells(
    sheet_type: str, rows: list, cell_rows: list
) -> tuple[list, str]:
    """
    Parse a legacy print-style TPA loss run sheet.

    Layout quirks handled:
    - Two-row column headers (group label row + sub-label row)
    - Address / cause-of-loss sub-rows interleaved with claim rows
    - '----------' separator rows between sections
    - 'Total …' section subtotal rows (skipped as data, captured as totals)
    """
    header_pair = _find_legacy_header_rows(rows)
    if header_pair is None:
        # Fall back to standard parser
        hri = _find_header_row(rows)
        if hri is None:
            return [], sheet_type
        headers = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        data_start = hri + 1
    else:
        top_hri, bot_hri = header_pair
        if top_hri == bot_hri:
            # Single-row header variant
            headers = [
                str(h).strip() if h is not None else f"Column_{i}"
                for i, h in enumerate(rows[top_hri])
            ]
        else:
            headers = _merge_two_header_rows(rows[top_hri], rows[bot_hri])
        data_start = bot_hri + 1

    # Pad/trim headers to actual column count
    num_cols = max(len(rows[i]) for i in range(len(rows))) if rows else len(headers)
    while len(headers) < num_cols:
        headers.append(f"Column_{len(headers) + 1}")

    extracted: list[dict] = []
    pending_claim: dict | None = None   # last claim row waiting for sub-row

    for r_idx_rel, (raw_row, cell_row) in enumerate(
        zip(rows[data_start:], cell_rows[data_start:])
    ):
        r_idx = data_start + 1 + r_idx_rel  # 1-based Excel row number

        # --- Completely empty row: flush pending claim -------------------------
        if not any(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        # --- Separator row (----------): flush pending, skip ------------------
        if _is_separator_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        # --- Section subtotal row: flush pending, skip -----------------------
        if _is_subtotal_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        # --- Legacy sub-row (address + cause): enrich pending claim ----------
        if _is_legacy_sub_row(raw_row, num_cols):
            if pending_claim is not None:
                # Absorb address from col B (index 1)
                if len(raw_row) > 1 and raw_row[1] is not None:
                    addr_val = str(raw_row[1]).strip()
                    if addr_val:
                        _enrich_field(pending_claim, "Address", addr_val, r_idx, 2)
                # Absorb cause-of-loss from col D (index 3)
                if len(raw_row) > 3 and raw_row[3] is not None:
                    col_val = str(raw_row[3]).strip()
                    if col_val:
                        _enrich_field(pending_claim, "Cause of Loss", col_val, r_idx, 4)
            # Sub-row does NOT flush pending — the next claim row will flush it
            continue

        # --- Aggregate/totals heuristic (non-subtotal) -----------------------
        if _is_aggregate_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        # --- Normal claim row ------------------------------------------------
        # Flush any previously pending claim before starting a new one
        if pending_claim is not None:
            extracted.append(pending_claim)
            pending_claim = None

        row_data: dict = {}
        for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
            if c_idx_0 >= len(headers):
                continue
            header = headers[c_idx_0]
            if not header:
                continue
            clean_val = format_cell_value_with_fmt(cell)
            real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
            row_data[header] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": real_col,
            }

        if any(v["value"] for v in row_data.values()):
            pending_claim = row_data

    # Flush last pending claim
    if pending_claim is not None:
        extracted.append(pending_claim)

    return extracted, sheet_type


def _enrich_field(
    claim: dict, field_name: str, value: str, excel_row: int, excel_col: int
) -> None:
    """Add or update a field in a claim dict if not already set."""
    if field_name not in claim or not claim[field_name].get("value"):
        claim[field_name] = {
            "value": value, "modified": value,
            "excel_row": excel_row, "excel_col": excel_col,
        }


# ── CSV / plain parse_rows (no cell objects) ──────────────────────────────────

def parse_rows(sheet_type: str, rows: list) -> tuple[list, str]:
    if sheet_type == "SUMMARY":
        hri = None
        for i, row in enumerate(rows[:20]):
            rt = " ".join([str(c).lower() for c in row if c])
            if "sheet" in rt and "line of business" in rt:
                hri = i
                break
        if hri is None:
            return [], sheet_type
        headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        extracted = []
        for r_idx, row in enumerate(rows[hri + 1:], start=hri + 2):
            if not any(row):
                continue
            if _is_aggregate_row(list(row)):
                continue
            row_data: dict = {}
            for c_idx, value in enumerate(row, start=1):
                if c_idx - 1 >= len(headers):
                    continue
                clean_val = str(value).strip() if value is not None else ""
                row_data[headers[c_idx - 1]] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": c_idx,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    # Legacy check for CSV too
    if _is_legacy_print_layout(rows):
        # Build mock cell_rows (no real cell objects, reuse raw values)
        return _parse_legacy_layout_plain(sheet_type, rows)

    hri = _find_header_row(rows)
    if hri is None:
        return [], sheet_type
    headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
    extracted = []
    for r_idx, row in enumerate(rows[hri + 1:], start=hri + 2):
        if not any(row):
            continue
        if any(str(cell).lower().strip() in ["totals", "total", "grand total"] for cell in row if cell):
            break
        if _is_aggregate_row(list(row)):
            continue
        row_data: dict = {}
        for c_idx, value in enumerate(row, start=1):
            if c_idx - 1 >= len(headers):
                continue
            clean_val = str(value).strip() if value is not None else ""
            row_data[headers[c_idx - 1]] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


def _parse_legacy_layout_plain(sheet_type: str, rows: list) -> tuple[list, str]:
    """parse_rows equivalent for legacy layout when no cell objects are available."""
    header_pair = _find_legacy_header_rows(rows)
    if header_pair is None:
        hri = _find_header_row(rows)
        if hri is None:
            return [], sheet_type
        headers    = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        data_start = hri + 1
    else:
        top_hri, bot_hri = header_pair
        if top_hri == bot_hri:
            headers = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[top_hri])]
        else:
            headers = _merge_two_header_rows(rows[top_hri], rows[bot_hri])
        data_start = bot_hri + 1

    num_cols = max(len(r) for r in rows) if rows else len(headers)
    while len(headers) < num_cols:
        headers.append(f"Column_{len(headers) + 1}")

    extracted: list[dict] = []
    pending_claim: dict | None = None

    for r_idx, raw_row in enumerate(rows[data_start:], start=data_start + 1):
        if not any(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue
        if _is_separator_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue
        if _is_subtotal_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue
        if _is_legacy_sub_row(raw_row, num_cols):
            if pending_claim is not None:
                if len(raw_row) > 1 and raw_row[1] is not None:
                    _enrich_field(pending_claim, "Address", str(raw_row[1]).strip(), r_idx, 2)
                if len(raw_row) > 3 and raw_row[3] is not None:
                    _enrich_field(pending_claim, "Cause of Loss", str(raw_row[3]).strip(), r_idx, 4)
            continue
        if _is_aggregate_row(list(raw_row)):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if pending_claim is not None:
            extracted.append(pending_claim)
            pending_claim = None

        row_data: dict = {}
        for c_idx, value in enumerate(raw_row, start=1):
            if c_idx - 1 >= len(headers):
                continue
            header    = headers[c_idx - 1]
            if not header:
                continue
            clean_val = str(value).strip() if value is not None else ""
            row_data[header] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            pending_claim = row_data

    if pending_claim is not None:
        extracted.append(pending_claim)

    return extracted, sheet_type