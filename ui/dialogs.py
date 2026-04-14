"""
ui/dialogs.py
All @st.dialog popups:
  - show_eye_popup         — cell-view with Excel highlight
  - show_field_history_dialog — per-field edit timeline
  - show_settings_dialog   — conf threshold + schema manager
  - show_schema_fields_dialog — required / accepted / custom field viewer

AUDIT LOG BEHAVIOUR (show_claim_journey_dialog)
────────────────────────────────────────────────
- Default view  : only user actions THIS SESSION — FIELD_EDITED, FIELD_ADDED,
                  EXPORT_GENERATED with timestamp >= _session_start
- "View Full History" button : expands inline, ALL events across all sessions
- LLM_CAUSE_ENRICHED : deduplicated to FIRST occurrence per claim in full history;
                        never shown in the default user-actions view
- Each row has ▼/▲ toggle for full event-dict detail
- All toggles use on_click callbacks — never st.rerun() — so the dialog
  stays open when the user expands/collapses rows or switches history views

DIALOG PERSISTENCE
──────────────────
- show_claim_journey_dialog is kept open across reruns via the
  "_open_journey_dialog" session-state flag set by the caller (e.g. the
  "View Journey" button in claim_panel / export_panel).
- The flag is checked at the top of app.py on every rerun, so on_click
  callbacks inside the dialog (which trigger a rerun) automatically
  re-open the dialog in the correct state.
- Only the Close button clears the flag, which is the one place we
  actually want the dialog to disappear.
- Callers must NOT call show_claim_journey_dialog() directly from a
  button handler.  Instead set the flag:
      st.session_state["_open_journey_dialog"] = {
          "claim_id": ..., "curr_claim": ...,
          "selected_sheet": ..., "active_schema": ...,
          "_llm_map_result": ...,
      }
  app.py will call show_claim_journey_dialog() on the next rerun.
"""

import csv
import datetime

import streamlit as st
from openpyxl.utils import get_column_letter
from PIL import ImageDraw

from modules.field_history import _get_field_history
from modules.excel_renderer import (
    render_excel_sheet, get_cell_pixel_bbox, crop_context,
)


# ── Eye popup ─────────────────────────────────────────────────────────────────

@st.dialog("Cell View", width="large")
def show_eye_popup(field: str, info: dict, excel_path: str, sheet_name: str) -> None:
    import os
    raw_value  = info.get("value", "") or ""
    mod_value  = info.get("modified", raw_value) or raw_value
    target_row = info.get("excel_row")
    target_col = info.get("excel_col")
 
    st.markdown(f"### 📍 {field}")
 
    def _val_box(label: str, val: str, color: str = "#4f9cf9"):
        _empty_html = "<span style='color:#555;'>( empty )</span>"
        _content    = val if val else _empty_html
        st.markdown(
            f"<div style='margin-bottom:12px;'>"
            f"<div style='font-size:10px;font-weight:700;color:{color};font-family:monospace;"
            f"text-transform:uppercase;letter-spacing:1.2px;margin-bottom:5px;'>{label}</div>"
            f"<div style='background:#12121c;border:1px solid #2a2a45;border-radius:6px;"
            f"padding:12px 14px;font-family:Consolas,monospace;font-size:13px;"
            f"color:#e8e7ff;word-break:break-all;white-space:pre-wrap;"
            f"max-height:200px;overflow-y:auto;line-height:1.6;'>"
            f"{_content}"
            f"</div></div>",
            unsafe_allow_html=True,
        )
 
    _val_box("Extracted Value (raw from file)", raw_value, "#34d399")
    if mod_value and mod_value != raw_value:
        _val_box("Modified Value (user edited)", mod_value, "#f5c842")
 
    ext = os.path.splitext(excel_path)[1].lower()
 
    if target_row and target_col:
        col_letter = get_column_letter(target_col)
        st.markdown(
            f"<div style='font-size:12px;color:#a0a0c8;font-family:monospace;margin-bottom:12px;'>"
            f"📌 Cell <b style='color:#4f9cf9;'>{col_letter}{target_row}</b>"
            f" &nbsp;·&nbsp; Row {target_row} &nbsp;·&nbsp; Col {target_col}"
            f"</div>",
            unsafe_allow_html=True,
        )
    elif target_row and ext == ".pdf":
        st.markdown(
            f"<div style='font-size:12px;color:#a0a0c8;font-family:monospace;margin-bottom:12px;'>"
            f"📄 PDF Page <b style='color:#4f9cf9;'>{target_row}</b>"
            f"</div>",
            unsafe_allow_html=True,
        )
    else:
        st.warning("No cell location recorded for this field.")
        return
 
    st.markdown("---")
 
    # ── PDF branch ────────────────────────────────────────────────────────────
    if ext == ".pdf":
        source_text      = info.get("source_text", "")
        bounding_polygon = info.get("bounding_polygon")
        page_width       = info.get("page_width")  or 8.5
        page_height      = info.get("page_height") or 11.0
 
        st.markdown(
            f"<div style='font-size:10px;color:#4f9cf9;font-weight:700;font-family:monospace;"
            f"text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;'>"
            f"📄 PDF Source — Page {target_row}</div>",
            unsafe_allow_html=True,
        )
 
        if bounding_polygon:
            # ── Render PDF page with Azure DI bounding box highlight ──────────
            _pdf_cache_key = f"_pdf_render_{excel_path}_{target_row}_{field}"
            with st.spinner("Rendering PDF page…"):
                if _pdf_cache_key not in st.session_state:
                    try:
                        from modules.excel_renderer import render_pdf_page_with_highlight
                        full_img, cropped_img = render_pdf_page_with_highlight(
                            pdf_path           = excel_path,
                            page_number        = int(target_row),
                            bounding_polygon   = bounding_polygon,
                            page_width_inches  = float(page_width),
                            page_height_inches = float(page_height),
                            dpi                = 150,
                        )
                        st.session_state[_pdf_cache_key] = (full_img, cropped_img)
                    except Exception as e:
                        st.session_state[_pdf_cache_key] = (None, None)
                        st.error(f"PDF render error: {e}")
 
                full_img, cropped_img = st.session_state.get(_pdf_cache_key, (None, None))
 
            if cropped_img is not None:
                st.image(
                    cropped_img,
                    use_container_width=True,
                    caption=f"Field '{field}' highlighted on PDF Page {target_row}",
                )
                # Show source text below the image as context
                if source_text:
                    st.markdown(
                        f"<div style='background:#12121c;border:1px solid #2a2a45;"
                        f"border-radius:6px;padding:8px 12px;font-family:monospace;"
                        f"font-size:11px;color:#a0a0c8;margin-top:8px;'>"
                        f"{source_text}</div>",
                        unsafe_allow_html=True,
                    )
            else:
                # Render failed — fall back to source text
                st.warning(
                    "Could not render PDF page image. "
                    "Install pymupdf with: `pip install pymupdf`"
                )
                st.markdown(
                    f"<div style='background:#12121c;border:1px solid #2a2a45;"
                    f"border-radius:6px;padding:12px 14px;font-family:monospace;"
                    f"font-size:12px;color:#a0a0c8;'>"
                    f"{source_text or '(no source text recorded)'}</div>",
                    unsafe_allow_html=True,
                )
        else:
            # ── No bounding polygon — use text-search highlight fallback ──────
            _pdf_cache_key = f"_pdf_render_{excel_path}_{target_row}_{field}_textfallback"
            with st.spinner("Rendering PDF page…"):
                if _pdf_cache_key not in st.session_state:
                    try:
                        from modules.excel_renderer import render_pdf_page_text_highlight
                        full_img, cropped_img = render_pdf_page_text_highlight(
                            pdf_path    = excel_path,
                            page_number = int(target_row),
                            search_text = raw_value or source_text or field,
                            dpi         = 150,
                        )
                        st.session_state[_pdf_cache_key] = (full_img, cropped_img)
                    except Exception as e:
                        st.session_state[_pdf_cache_key] = (None, None)

                full_img, cropped_img = st.session_state.get(_pdf_cache_key, (None, None))

            if cropped_img is not None:
                st.image(
                    cropped_img,
                    use_container_width=True,
                    caption=f"Field '{field}' highlighted on PDF Page {target_row}",
                )
            else:
                st.markdown(
                    f"<div style='background:#12121c;border:1px solid #2a2a45;border-radius:6px;"
                    f"padding:12px 14px;font-family:monospace;font-size:12px;color:#a0a0c8;"
                    f"line-height:1.6;margin-bottom:8px;'>"
                    f"{source_text or '(no source text recorded)'}"
                    f"</div>",
                    unsafe_allow_html=True,
                )

            if source_text:
                st.markdown(
                    f"<div style='background:#12121c;border:1px solid #2a2a45;"
                    f"border-radius:6px;padding:8px 12px;font-family:monospace;"
                    f"font-size:11px;color:#a0a0c8;margin-top:8px;'>"
                    f"{source_text}</div>",
                    unsafe_allow_html=True,
                )
            st.info(
                "📝 This field was extracted from page text — highlighted by text search."
            )
        return
 
    # ── CSV branch ────────────────────────────────────────────────────────────
    if ext == ".csv":
        try:
            import csv as _csv
            with open(excel_path, "r", encoding="utf-8-sig") as f:
                all_rows = list(_csv.reader(f))
            if not all_rows:
                return
            n_rows = len(all_rows)
            n_cols = max(len(r) for r in all_rows)
            r0, r1 = max(0, target_row - 4), min(n_rows, target_row + 4)
 
            col_headers = "".join(
                f"<th style='background:#1a1a2e;color:#6b7280;font-size:11px;"
                f"padding:5px 10px;border:1px solid #2a2a45;font-family:monospace;"
                f"font-weight:600;text-align:center;'>{get_column_letter(c+1)}</th>"
                for c in range(n_cols)
            )
            thead = (
                f"<thead><tr>"
                f"<th style='background:#1a1a2e;color:#6b7280;font-size:11px;"
                f"padding:5px 8px;border:1px solid #2a2a45;font-family:monospace;'>#</th>"
                f"{col_headers}</tr></thead>"
            )
            tbody = ""
            for r_idx in range(r0, r1):
                row_data = all_rows[r_idx] if r_idx < len(all_rows) else []
                is_tr    = (r_idx + 1 == target_row)
                rn_bg    = "#1a2540" if is_tr else "#12121c"
                rn_color = "#4f9cf9" if is_tr else "#555"
                cells = (
                    f"<td style='background:{rn_bg};color:{rn_color};font-size:11px;"
                    f"padding:5px 8px;border:1px solid #2a2a45;font-family:monospace;"
                    f"font-weight:bold;text-align:center;'>{r_idx+1}</td>"
                )
                for c_idx in range(n_cols):
                    cell_val = row_data[c_idx] if c_idx < len(row_data) else ""
                    is_tc    = is_tr and (c_idx + 1 == target_col)
                    if is_tc:
                        style = "background:#2a2010;border:2px solid #f5c842;color:#fff;font-weight:bold;"
                    elif is_tr:
                        style = "background:#1a2540;border:1px solid rgba(79,156,249,0.3);color:#c8d8ff;"
                    else:
                        style = "background:#12121c;border:1px solid #2a2a45;color:#6b7280;"
                    cells += (
                        f"<td style='{style}font-size:11px;padding:5px 10px;"
                        f"white-space:normal;word-break:break-word;"
                        f"font-family:monospace;'>{cell_val}</td>"
                    )
                tbody += f"<tr>{cells}</tr>"
 
            st.markdown(
                f"<div style='overflow-x:auto;border-radius:6px;border:1px solid #2a2a45;'>"
                f"<table style='border-collapse:collapse;width:100%;'>"
                f"{thead}<tbody>{tbody}</tbody></table></div>",
                unsafe_allow_html=True,
            )
        except Exception as e:
            st.error(f"CSV preview error: {e}")
        return
 
    # ── Excel branch ──────────────────────────────────────────────────────────
    cache_key = f"_rendered_{excel_path}_{sheet_name}"
    with st.spinner("Rendering sheet…"):
        if cache_key not in st.session_state:
            rendered_img, col_starts, row_starts, merged_master = render_excel_sheet(
                excel_path, sheet_name, scale=1.0
            )
            st.session_state[cache_key] = (rendered_img, col_starts, row_starts, merged_master)
        else:
            rendered_img, col_starts, row_starts, merged_master = st.session_state[cache_key]
 
    try:
        img  = rendered_img.copy()
        draw = ImageDraw.Draw(img, "RGBA")
        x1, y1, x2, y2 = get_cell_pixel_bbox(col_starts, row_starts, target_row, target_col, merged_master)
        draw.rectangle([x1+1, y1+1, x2-1, y2-1], fill=(255, 230, 0, 80))
        draw.rectangle([x1, y1, x2, y2], outline=(245, 158, 11, 255), width=3)
        draw.rectangle([x1+3, y1+3, x2-3, y2-3], outline=(255, 255, 255, 160), width=1)
        cropped, _, _, _, _ = crop_context(img, x1, y1, x2, y2, pad_x=300, pad_y=200)
        col_letter = get_column_letter(target_col)
        st.image(cropped, use_container_width=True,
                 caption=f"Cell {col_letter}{target_row} highlighted in yellow")
    except Exception as e:
        st.error(f"Rendering error: {e}")


# ── Field history dialog ──────────────────────────────────────────────────────

@st.dialog("Field History", width="large")
def show_field_history_dialog(
    field_name: str, sheet: str, claim_id: str,
    current_val: str, original_val: str,
) -> None:
    """
    Render a modal popup showing the full edit timeline for a single field.

    Displays a two-column header comparing the original extracted value
    against the current value, with a colour-coded indicator showing
    whether the field has been modified. Below the header, each recorded
    edit is shown as a timestamped row with source icon (manual edit vs
    auto/LLM), a FROM → TO value diff, and a bottom border separator.

    Args:
        field_name (str): The standard field name shown in the dialog
            heading (e.g. "Total Paid", "Status").
        sheet (str): Active sheet/tab name. Used to scope the history
            key lookup in :func:`modules.field_history._get_field_history`.
        claim_id (str): Stable claim identifier. Used alongside ``sheet``
            and ``field_name`` to look up the correct history list.
        current_val (str): The current displayed value for the field
            (may be the original, a normalised value, or a user edit).
        original_val (str): The raw value extracted directly from the
            source Excel or CSV file, shown as the "before" baseline.

    Returns:
        None. All output is written directly to the Streamlit dialog.

    Side effects:
        - Reads ``st.session_state`` via
          :func:`modules.field_history._get_field_history`.
        - Calls ``st.rerun()`` when the Close button is clicked to
          dismiss the dialog.

    Example trigger:
        >>> # Called when user clicks the history icon on a field row
        >>> show_field_history_dialog(
        ...     "Total Paid", "Q1 Claims", "CLM-001",
        ...     current_val="5000.00", original_val="$5,000"
        ... )

    Dependencies:
        - :func:`modules.field_history._get_field_history` : loads the
          edit history list for the field from session state.
        - ``streamlit``                                    : dialog rendering.
    """
    st.markdown(f"### 📋 History — {field_name}")
    history = _get_field_history(sheet, claim_id, field_name)

    st.markdown(
        f"""
        <div style='background:var(--s0);border:1px solid var(--b0);border-radius:8px;
             padding:12px 16px;margin-bottom:12px;'>
          <div style='display:grid;grid-template-columns:1fr 1fr;gap:16px;'>
            <div>
              <div style='font-size:10px;color:var(--t3);font-family:monospace;text-transform:uppercase;
                   letter-spacing:1px;margin-bottom:6px;'>Original (from file)</div>
              <div style='background:#1a1a2e;border:1px solid #2a2a45;border-radius:5px;
                   padding:8px 12px;font-family:monospace;font-size:13px;color:#f0efff;'>
                {original_val or "(empty)"}
              </div>
            </div>
            <div>
              <div style='font-size:10px;color:var(--t3);font-family:monospace;text-transform:uppercase;
                   letter-spacing:1px;margin-bottom:6px;'>Current Value</div>
              <div style='background:#0f2d1f;border:1px solid rgba(52,211,153,0.35);border-radius:5px;
                   padding:8px 12px;font-family:monospace;font-size:13px;color:#34d399;'>
                {current_val or "(empty)"}
              </div>
            </div>
          </div>
          {"<div style='margin-top:8px;font-size:11px;color:#f5c842;font-family:monospace;'>⚡ Modified from original</div>" if current_val != original_val else "<div style='margin-top:8px;font-size:11px;color:#34d399;font-family:monospace;'>✓ Unchanged from original</div>"}
        </div>
        """,
        unsafe_allow_html=True,
    )

    if history:
        st.markdown("**Edit Timeline**")
        for h in history:
            arrow_col = "var(--yellow)" if h["source"] == "user" else "var(--blue)"
            src_icon  = "✏" if h["source"] == "user" else "⚡"
            src_lbl   = "Manual edit" if h["source"] == "user" else "Auto (LLM/normalize)"
            st.markdown(
                f"<div style='display:flex;align-items:flex-start;gap:12px;padding:10px 0;"
                f"border-bottom:1px solid #1e1e32;'>"
                f"<div style='font-size:10px;color:var(--t3);font-family:monospace;"
                f"white-space:nowrap;margin-top:2px;'>{h['ts']}</div>"
                f"<div style='flex:1;'>"
                f"<div style='display:flex;align-items:center;gap:8px;margin-bottom:6px;'>"
                f"<span style='color:{arrow_col};font-size:12px;'>{src_icon}</span>"
                f"<span style='font-size:11px;color:var(--t3);font-family:monospace;'>{src_lbl}</span></div>"
                f"<div style='display:flex;align-items:center;gap:8px;'>"
                f"<code style='background:#1a1a2e;padding:3px 8px;border-radius:4px;font-size:12px;color:#f0efff;'>"
                f"{h['from'] or '(empty)'}</code>"
                f"<span style='color:{arrow_col};font-size:14px;'>→</span>"
                f"<code style='background:#0f2d1f;padding:3px 8px;border-radius:4px;font-size:12px;color:#34d399;'>"
                f"{h['to'] or '(empty)'}</code></div></div></div>",
                unsafe_allow_html=True,
            )
    else:
        st.markdown(
            "<div style='color:var(--t3);font-size:13px;padding:12px 0;'>"
            "No edits recorded yet for this field.</div>",
            unsafe_allow_html=True,
        )

    if st.button("Close", type="primary", use_container_width=True):
        st.rerun()


# ── Settings dialog ───────────────────────────────────────────────────────────

@st.dialog("Settings", width="large")
def show_settings_dialog(schemas: dict, config_load_status: dict) -> None:
    """
    Render the application Settings modal with three sections:

    1. Confidence Settings — toggle confidence display on/off and set
       the numeric threshold (0–100) via a slider. The slider renders
       a colour-coded progress bar and a plain-English description of
       the chosen level.

    2. Export Schema — lists every available schema (Standard, Guidewire,
       Duck Creek, etc.) with Activate/Deactivate, View Fields, and
       Custom Fields buttons. The active schema is visually distinguished
       with a coloured border and an "● ACTIVE" badge.

    3. YAML Config Files — shows the load status of each schema's YAML
       config file with a green "✓ Loaded" or red "✗ Not found" badge,
       the file path, and a "Reload YAML Configs" button that hot-reloads
       all schemas without requiring an app restart.

    A "Reset Defaults" button in the footer restores the confidence
    threshold to 80, disables confidence display, clears the active
    schema, and removes all custom fields for every schema.

    Args:
        schemas (dict): The full ``SCHEMAS`` dict from ``config.schemas``,
            mapping schema names to their definition dicts (icon, version,
            description, color, required_fields, accepted_fields, etc.).
        config_load_status (dict): Maps schema names to a status dict::

                {
                    "loaded": bool,   # True if the YAML file was found
                    "file":   str,    # filename of the YAML config
                }

    Returns:
        None. All output is written directly to the Streamlit dialog.

    Side effects:
        - Reads and writes ``st.session_state`` keys:
          "use_conf_threshold", "conf_threshold", "active_schema",
          ``f"custom_fields_{schema_name}"`` for each schema,
          "sheet_cache", "schema_popup_target", "schema_popup_tab".
        - Calls ``st.rerun()`` on every button click (Activate, View
          Fields, Custom Fields, Reset Defaults, Reload YAML, Close).
        - On "Reload YAML Configs", replaces ``config.schemas.SCHEMAS``
          in-place and clears "sheet_cache" to force re-parsing.

    Example trigger:
        >>> # Called from the ⚙ Settings button in the main toolbar
        >>> show_settings_dialog(schemas=SCHEMAS, config_load_status=status)

    Dependencies:
        - ``config.schemas._load_all_configs``  : YAML hot-reload.
        - ``config.settings.CONFIG_DIR``        : displayed in the YAML section.
        - ``streamlit``                         : dialog, widgets, and session state.
    """
    import os
    from config.settings import CONFIG_DIR

    st.markdown("### Configuration")
    st.markdown("---")
    st.markdown("#### Confidence Settings")

    use_conf = st.checkbox(
        "Enable confidence scoring display",
        value=st.session_state.get("use_conf_threshold", False),
        key="use_conf_toggle",
        help="When enabled, shows confidence scores for each mapped field",
    )
    st.session_state["use_conf_threshold"] = use_conf

    if use_conf:
        conf = st.slider(
            "Confidence threshold", 0, 100,
            value=st.session_state.get("conf_threshold", 80),
            step=5, format="%d%%",
        )
        st.session_state["conf_threshold"] = conf
        bar_color = "#22c55e" if conf >= 70 else "#f59e0b" if conf >= 40 else "#ef4444"
        level_txt = (
            "High confidence — minimal manual review needed" if conf >= 70 else
            "Medium — review flagged fields carefully" if conf >= 40 else
            "Low — most fields will require manual review"
        )
        st.markdown(
            f"<div class='conf-bar-wrap'><div class='conf-bar-fill' "
            f"style='width:{conf}%;background:{bar_color};'></div></div>"
            f"<div style='color:{bar_color};font-size:12px;margin-top:5px;'>{level_txt}</div>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            "<div style='color:var(--t3);font-size:13px;font-family:monospace;'>"
            "Confidence scoring is disabled. Enable above to show scores and set threshold.</div>",
            unsafe_allow_html=True,
        )

    st.markdown("---")
    st.markdown("#### Export Schema")
    active_schema = st.session_state.get("active_schema", None)

    for schema_name, schema_def in schemas.items():
        is_active   = active_schema == schema_name
        border_col  = schema_def["color"] if is_active else "#2a2a3e"
        bg_col      = "#1a1a2e" if is_active else "#16161e"
        active_tag  = (
            f"<span style='font-size:10px;color:{schema_def['color']};margin-left:8px;font-weight:bold;'>● ACTIVE</span>"
            if is_active else ""
        )
        custom_count = len(st.session_state.get(f"custom_fields_{schema_name}", []))
        st.markdown(
            f"<div style='background:{bg_col};border:1px solid {border_col};border-radius:8px;"
            f"padding:12px 14px;margin-bottom:4px;'>"
            f"<div style='display:flex;align-items:center;'>"
            f"<span style='font-size:var(--sz-body);font-weight:700;color:var(--t0);font-family:var(--font);'>"
            f"{schema_def['icon']} {schema_name}</span>"
            f"<span style='font-size:var(--sz-sm);color:var(--t3);margin-left:8px;font-family:var(--font);'>"
            f"{schema_def['version']}</span>{active_tag}</div>"
            f"<div style='font-size:var(--sz-sm);color:var(--t2);margin-top:4px;font-family:var(--font);'>"
            f"{schema_def['description']}</div></div>",
            unsafe_allow_html=True,
        )
        bc1, bc2, bc3 = st.columns([1, 1, 1])
        with bc1:
            if st.button(
                "✓ Deactivate" if is_active else "Activate",
                key=f"activate_{schema_name}", use_container_width=True,
            ):
                st.session_state["active_schema"] = None if is_active else schema_name
                st.rerun()
        with bc2:
            if st.button("View Fields", key=f"view_{schema_name}", use_container_width=True):
                st.session_state["schema_popup_target"] = schema_name
                st.session_state["schema_popup_tab"]    = "required"
                st.rerun()
        with bc3:
            if st.button(
                f"Custom Fields ({custom_count})",
                key=f"custom_{schema_name}", use_container_width=True,
            ):
                st.session_state["schema_popup_target"] = schema_name
                st.session_state["schema_popup_tab"]    = "custom"
                st.rerun()
        st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("#### 📁 YAML Config Files")
    st.markdown(
        f"<div style='color:var(--t2);font-size:var(--sz-body);margin-bottom:10px;font-family:var(--font);'>"
        f"Config directory: <code>{CONFIG_DIR}</code></div>",
        unsafe_allow_html=True,
    )
    for schema_name, status in config_load_status.items():
        sc      = schemas.get(schema_name, {})
        col_st  = sc.get("color", "#64748b")
        badge   = (
            "<span style='background:#0f2d1f;border:1px solid #22c55e;border-radius:4px;"
            "padding:1px 7px;font-size:10px;color:#22c55e;'>✓ Loaded</span>"
            if status["loaded"]
            else
            "<span style='background:#2d0f0f;border:1px solid #ef4444;border-radius:4px;"
            "padding:1px 7px;font-size:10px;color:#ef4444;'>✗ Not found — using defaults</span>"
        )
        st.markdown(
            f"<div style='background:var(--s0);border:1px solid var(--b0);border-radius:6px;"
            f"padding:10px 14px;margin-bottom:6px;'>"
            f"<div style='display:flex;align-items:center;gap:10px;'>"
            f"<span style='color:{col_st};font-weight:700;font-size:var(--sz-body);font-family:var(--font);'>"
            f"{sc.get('icon','')} {schema_name}</span>{badge}</div>"
            f"<div style='font-size:var(--sz-xs);color:var(--t3);margin-top:4px;font-family:var(--font);'>"
            f"📄 {status['file']}</div></div>",
            unsafe_allow_html=True,
        )

    if st.button("🔄 Reload YAML Configs", use_container_width=True, key="reload_yaml_cfg"):
        from config.schemas import _load_all_configs, _HARDCODED_SCHEMAS
        import config.schemas as _cs
        _cs.SCHEMAS = _load_all_configs(_HARDCODED_SCHEMAS)
        st.session_state["sheet_cache"] = {}
        st.success("✅ Configs reloaded")
        st.rerun()

    st.markdown("---")
    r1, r2 = st.columns(2)
    with r1:
        if st.button("Reset Defaults", use_container_width=True, key="reset_defaults_btn"):
            st.session_state["conf_threshold"]     = 80
            st.session_state["use_conf_threshold"] = False
            st.session_state["active_schema"]      = None
            for s in schemas:
                st.session_state[f"custom_fields_{s}"] = []
            st.rerun()
    with r2:
        if st.button("Close", type="primary", use_container_width=True):
            st.rerun()


# ── Schema field manager dialog ───────────────────────────────────────────────

@st.dialog("Schema Field Manager", width="large")
def show_schema_fields_dialog(schema_name: str, schemas: dict) -> None:
    """
    Render the Schema Field Manager modal with three tabs for browsing
    and customising the fields of a single schema.

    Tabs:
        - Mandatory Fields  : pill display of all required fields.
        - All Accepted Fields : pills split into MANDATORY and OPTIONAL
          groups, showing every field the schema recognises.
        - My Custom Fields  : interactive panel to add optional fields
          from the schema's accepted list and remove them individually
          or all at once. Shows a live summary of mandatory / custom /
          total field counts.

    Custom fields are stored in
    ``st.session_state[f"custom_fields_{schema_name}"]`` as a list.
    Adding a field appends to this list; removing pops by index. The
    selectbox for adding is populated only with fields not already
    present in the required or custom lists (one-to-one constraint).

    Args:
        schema_name (str): Key of the schema to manage
            (e.g. "Standard", "Guidewire"). Must exist in ``schemas``.
        schemas (dict): Full ``SCHEMAS`` dict from ``config.schemas``,
            mapping schema names to their definition dicts. Must contain
            at minimum for the given ``schema_name``:
            - "icon"            (str)       : emoji icon.
            - "version"         (str)       : version string.
            - "description"     (str)       : one-line description.
            - "required_fields" (list[str]) : mandatory field names.
            - "accepted_fields" (list[str]) : all recognised field names.

    Returns:
        None. All output is written directly to the Streamlit dialog.

    Side effects:
        - Initialises ``st.session_state[f"custom_fields_{schema_name}"]``
          to ``[]`` if not already present.
        - Mutates the custom fields list on Add / Remove / Clear All,
          followed by ``st.rerun()``.

    Example trigger:
        >>> # Called when user clicks "View Fields" in the Settings dialog
        >>> show_schema_fields_dialog("Guidewire", schemas=SCHEMAS)

    Dependencies:
        - ``streamlit`` : tabs, selectbox, buttons, and session state.
    """
    schema     = schemas[schema_name]
    custom_key = f"custom_fields_{schema_name}"
    if custom_key not in st.session_state:
        st.session_state[custom_key] = []

    st.markdown(f"### {schema['icon']} {schema_name} — {schema['version']}")
    st.markdown(
        f"<div style='color:var(--t2);font-size:var(--sz-body);margin-bottom:14px;font-family:var(--font);'>"
        f"{schema['description']}</div>",
        unsafe_allow_html=True,
    )
    tab_req, tab_accepted, tab_custom = st.tabs(["Mandatory Fields", "All Accepted Fields", "My Custom Fields"])

    with tab_req:
        pills = "".join(
            f"<span class='field-pill field-pill-required'>✓ {f}</span>"
            for f in schema["required_fields"]
        )
        st.markdown(f"<div style='margin:12px 0;'>{pills}</div>", unsafe_allow_html=True)

    with tab_accepted:
        optional  = [f for f in schema["accepted_fields"] if f not in schema["required_fields"]]
        req_pills = "".join(f"<span class='field-pill field-pill-required'>✓ {f}</span>" for f in schema["required_fields"])
        opt_pills = "".join(f"<span class='field-pill'>{f}</span>" for f in optional)
        st.markdown(
            f"<div style='margin:12px 0;'>"
            f"<b style='color:var(--t2);font-size:var(--sz-xs);font-family:var(--font);letter-spacing:1.2px;"
            f"text-transform:uppercase;'>MANDATORY</b>"
            f"<br><div style='margin-top:6px;'>{req_pills}</div></div>"
            f"<div style='margin:12px 0;'>"
            f"<b style='color:var(--t2);font-size:var(--sz-xs);font-family:var(--font);letter-spacing:1.2px;"
            f"text-transform:uppercase;'>OPTIONAL</b>"
            f"<br><div style='margin-top:6px;'>{opt_pills}</div></div>",
            unsafe_allow_html=True,
        )

    with tab_custom:
        custom_fields = st.session_state[custom_key]
        already_added = set(custom_fields) | set(schema["required_fields"])
        available     = [f for f in schema["accepted_fields"] if f not in already_added]

        if available:
            sel_col, add_col = st.columns([4, 1])
            with sel_col:
                chosen = st.selectbox(
                    "Pick field",
                    ["— select a field —"] + available,
                    key=f"new_field_sel_{schema_name}",
                    label_visibility="collapsed",
                )
            with add_col:
                if st.button("Add", key=f"add_field_btn_{schema_name}", use_container_width=True, type="primary"):
                    if chosen and chosen != "— select a field —":
                        st.session_state[custom_key].append(chosen)
                        st.rerun()

        if not custom_fields:
            st.markdown(
                "<div style='color:var(--t2);font-size:var(--sz-body);padding:10px 0;font-family:var(--font);'>"
                "No optional fields added yet.</div>",
                unsafe_allow_html=True,
            )
        else:
            for idx, cf in enumerate(list(custom_fields)):
                cf1, cf2 = st.columns([5, 1])
                with cf1:
                    cls = "field-pill-required" if cf in schema["required_fields"] else "field-pill-custom"
                    icon = "✓" if cf in schema["required_fields"] else "+"
                    st.markdown(f"<span class='field-pill {cls}'>{icon} {cf}</span>", unsafe_allow_html=True)
                with cf2:
                    if st.button("Remove", key=f"del_cf_{schema_name}_{idx}", use_container_width=True):
                        st.session_state[custom_key].pop(idx)
                        st.rerun()
            st.markdown("---")
            if st.button(f"Clear All", key=f"clear_all_{schema_name}"):
                st.session_state[custom_key] = []
                st.rerun()

        total = len(schema["required_fields"]) + len(custom_fields)
        st.markdown(
            f"<div style='background:var(--s0);border:1px solid var(--b0);border-radius:8px;"
            f"padding:10px 16px;'>"
            f"<span style='color:var(--t2);font-size:var(--sz-body);font-family:var(--font);'>"
            f"Mandatory: <b style='color:var(--blue);'>{len(schema['required_fields'])}</b>"
            f" &nbsp;|&nbsp; Custom: <b style='color:var(--green);'>{len(custom_fields)}</b>"
            f" &nbsp;|&nbsp; Total: <b style='color:var(--t0);'>{total}</b></span></div>",
            unsafe_allow_html=True,
        )


# ── Cache Manager dialog ──────────────────────────────────────────────────────

@st.dialog("Cache Manager", width="large")
def show_cache_manager_dialog() -> None:
    """
    Render the Cache Manager modal for inspecting and selectively
    clearing each cache layer used by the application.

    Displays live statistics for five cache layers (parsed sheet cache,
    file hash store, claim duplicate store, audit log, export history),
    then provides two clearing interfaces:

    Quick Presets:
        - Soft Reset        : clears UI session state only, preserving all
                              on-disk caches.
        - Clear File History: resets the hash store and claim duplicate
                              store so all files appear as new on next upload.
        - Full Reset        : clears all layers after a two-step
                              confirmation prompt.

    Granular Checkboxes:
        Allows independently selecting any combination of the six cache
        layers (UI state, parsed cache, file history, claim dups, audit
        log, export history) and clearing only those via "Clear Selected".

    The dialog never affects uploaded source files — only derived caches
    and logs are cleared.

    Args:
        None. All data is read from session state and on-disk cache files
        via the ``modules.cache_manager`` helpers.

    Returns:
        None. All output is written directly to the Streamlit dialog.

    Side effects:
        - Reads cache statistics via ``get_cache_stats()``.
        - On preset/granular clear: calls the appropriate
          ``modules.cache_manager`` clear functions and may set
          ``st.session_state["sheet_cache"] = {}`` to force re-parsing.
        - Sets ``st.session_state["_confirm_full_reset"] = True`` to
          trigger the confirmation prompt for Full Reset.
        - Calls ``st.rerun()`` after every mutating action.

    Example trigger:
        >>> # Called from the 🗄 Cache Manager button in the main toolbar
        >>> show_cache_manager_dialog()

    Dependencies:
        - ``modules.cache_manager.get_cache_stats``      : live stats.
        - ``modules.cache_manager.clear_session_cache``  : UI state clear.
        - ``modules.cache_manager.clear_parsed_cache``   : JSON cache clear.
        - ``modules.cache_manager.clear_hash_store``     : duplicate memory clear.
        - ``modules.cache_manager.clear_claim_dup_store``: claim dup clear.
        - ``modules.cache_manager.clear_audit_log``      : audit log clear.
        - ``modules.cache_manager.clear_export_table``   : export history clear.
        - ``modules.cache_manager._fmt_size``            : KB formatting helper.
        - ``streamlit``                                  : dialog and widgets.
    """
    from modules.cache_manager import (
        get_cache_stats, clear_session_cache, clear_parsed_cache,
        clear_hash_store, clear_claim_dup_store,
        clear_audit_log, clear_export_table, _fmt_size,
    )

    st.markdown("### 🗄️ Cache Manager")
    st.markdown(
        "<div style='font-size:13px;color:#a0a0c8;margin-bottom:16px;'>"
        "View and selectively clear each cache layer. "
        "This does not affect your uploaded files.</div>",
        unsafe_allow_html=True,
    )

    stats = get_cache_stats()

    def _stat_row(label, detail, color="#4f9cf9"):
        """
        Render a single cache-layer statistics row as a dark card.

        Displays the layer name on the left and a monospace detail string
        (e.g. "3 file(s) · 1.2 MB") on the right, with the detail
        coloured green/amber/red to indicate whether data is present.

        Args:
            label (str): Human-readable cache layer name
                (e.g. "Parsed Sheet Cache").
            detail (str): Pre-formatted detail string showing size or
                entry count (e.g. "3 file(s) · 1.2 MB").
            color (str): CSS hex colour for the detail text. Defaults
                to blue (#4f9cf9). Callers typically pass green when
                data is present and grey when the cache is empty.

        Returns:
            None. Writes directly to the Streamlit dialog via
            ``st.markdown(..., unsafe_allow_html=True)``.
        """
        st.markdown(
            f"<div style='display:flex;justify-content:space-between;align-items:center;"
            f"background:#17172a;border:1px solid #2a2a45;border-radius:6px;"
            f"padding:10px 14px;margin-bottom:6px;'>"
            f"<div style='font-size:13px;font-weight:600;color:#e8e7ff;'>{label}</div>"
            f"<div style='font-size:12px;font-family:monospace;color:{color};'>{detail}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )

    _stat_row("Parsed Sheet Cache",
              f"{stats['parsed']['files']} file(s) · {_fmt_size(stats['parsed']['size_kb'])}",
              "#34d399" if stats['parsed']['files'] > 0 else "#64748b")
    _stat_row("File Hash Store (Duplicate Memory)",
              f"{stats['hash_store']['entries']} file(s) tracked",
              "#f5c842" if stats['hash_store']['entries'] > 0 else "#64748b")
    _stat_row("Claim Duplicate Store",
              f"{stats['claim_dups']['entries']} claim(s) tracked",
              "#f87171" if stats['claim_dups']['entries'] > 0 else "#64748b")
    _stat_row("Audit Log",
              f"{stats['audit_log']['entries']} event(s) recorded",
              "#a78bfa" if stats['audit_log']['entries'] > 0 else "#64748b")
    _stat_row("Export History",
              f"{stats['export_table']['entries']} export(s) recorded",
              "#4f9cf9" if stats['export_table']['entries'] > 0 else "#64748b")

    st.markdown("---")
    st.markdown(
        "<div style='font-size:12px;color:#a0a0c8;margin-bottom:10px;font-family:monospace;"
        "text-transform:uppercase;letter-spacing:1px;'>Select what to clear</div>",
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns(2)
    with c1:
        do_session   = st.checkbox("UI Session State", value=True,
                                   help="Clears all in-memory selections, modified values, and panel states")
        do_parsed    = st.checkbox("Parsed Sheet Cache", value=False,
                                   help="Deletes cached JSON from feature_store/claims_json/ — next upload re-parses from scratch")
        do_hash      = st.checkbox("File Duplicate Memory", value=False,
                                   help="Resets hash_store.json — all files will be treated as NEW on next upload")
    with c2:
        do_claim_dup = st.checkbox("Claim Duplicate Store", value=False,
                                   help="Resets claim_dup_store.json — claim change tracking starts fresh")
        do_audit     = st.checkbox("Audit Log", value=False,
                                   help="Clears audit_log.json — all event history is lost")
        do_exports   = st.checkbox("Export History", value=False,
                                   help="Clears json_export_table.json — export records are removed")

    st.markdown("---")

    st.markdown(
        "<div style='font-size:12px;color:#a0a0c8;margin-bottom:8px;font-family:monospace;"
        "text-transform:uppercase;letter-spacing:1px;'>Quick presets</div>",
        unsafe_allow_html=True,
    )
    p1, p2, p3 = st.columns(3)
    with p1:
        if st.button("🔄 Soft Reset", use_container_width=True,
                     help="Clears UI state only — keeps all file history and cache on disk"):
            cleared = clear_session_cache(st.session_state)
            st.success(f"✅ UI state cleared ({cleared} keys removed)")
            st.rerun()
    with p2:
        if st.button("📁 Clear File History", use_container_width=True,
                     help="Resets duplicate memory so all files appear as NEW"):
            n_hash = clear_hash_store()
            n_dup  = clear_claim_dup_store()
            st.session_state["sheet_cache"] = {}
            st.success(f"✅ File history cleared ({n_hash} files, {n_dup} claims reset)")
            st.rerun()
    with p3:
        if st.button("🗑️ Full Reset", use_container_width=True,
                     help="Clears everything — session state, parsed cache, file history, claim tracking",
                     type="primary"):
            st.session_state["_confirm_full_reset"] = True
            st.rerun()

    if st.session_state.get("_confirm_full_reset"):
        st.warning("⚠️ **This will clear ALL cache layers.** Are you sure?")
        yes_col, no_col = st.columns(2)
        with yes_col:
            if st.button("Yes, clear everything", type="primary", use_container_width=True):
                clear_session_cache(st.session_state)
                clear_parsed_cache()
                clear_hash_store()
                clear_claim_dup_store()
                clear_audit_log()
                clear_export_table()
                st.session_state["_confirm_full_reset"] = False
                st.success("✅ All cache layers cleared. Upload a fresh file to begin.")
                st.rerun()
        with no_col:
            if st.button("Cancel", use_container_width=True):
                st.session_state["_confirm_full_reset"] = False
                st.rerun()

    st.markdown("---")

    col_clear, col_close = st.columns(2)
    with col_clear:
        if st.button("🗑️ Clear Selected", use_container_width=True):
            msgs = []
            if do_session:
                n = clear_session_cache(st.session_state)
                msgs.append(f"UI state ({n} keys)")
            if do_parsed:
                files, kb = clear_parsed_cache()
                msgs.append(f"Parsed cache ({files} files, {_fmt_size(kb)})")
                st.session_state["sheet_cache"] = {}
            if do_hash:
                n = clear_hash_store()
                msgs.append(f"File history ({n} entries)")
            if do_claim_dup:
                n = clear_claim_dup_store()
                msgs.append(f"Claim dups ({n} entries)")
            if do_audit:
                n = clear_audit_log()
                msgs.append(f"Audit log ({n} events)")
            if do_exports:
                n = clear_export_table()
                msgs.append(f"Export history ({n} entries)")
            if msgs:
                st.success("✅ Cleared: " + ", ".join(msgs))
                st.rerun()
            else:
                st.warning("Nothing selected — tick at least one checkbox above.")
    with col_close:
        if st.button("Close", type="primary", use_container_width=True):
            st.rerun()


# ── Claim Journey / Traceability Dialog ───────────────────────────────────────

@st.dialog("Claim Transformation Journey", width="large")
def show_claim_journey_dialog(
    claim_id: str,
    curr_claim: dict,
    selected_sheet: str,
    active_schema: str | None,
    _llm_map_result: dict,
) -> None:
    """
    Visual traceability view — full transformation journey for a claim.

    Renders a step-by-step breakdown of how every field in a claim was
    extracted, mapped, normalised, and edited, culminating in its current
    export-ready value.

    Sections rendered:
        1. Pipeline Trace banner — high-level ordered steps the claim
           passed through (FILE PARSED → SCHEMA MAPPED → LLM CALLED →
           USER EDITS) with timestamps and function references.
        2. Field Transformation Timeline — one card per field showing:
               Step 1 · Extracted from Document (source column, raw value).
               Step 2 · Mapping method (EXACT/FUZZY/PARTIAL/LLM/TITLE/DIRECT)
                        with header similarity score, value quality score,
                        and overall confidence percentage.
               Step N · User Edit(s) — one numbered step per recorded
                        edit showing FROM → TO with timestamp.
               Final  · Final Value card (only shown when field was edited).
        3. Audit Log section — session-filtered user-action events
           (FIELD_EDITED, FIELD_ADDED, EXPORT_GENERATED) with summary
           pills, and an optional "View Full History" inline expansion
           showing all events across all sessions.

    Audit log behaviour:
        - Default view  : only THIS SESSION's user actions, filtered by
          ``timestamp >= st.session_state["_session_start"]``.
        - Full history  : inline expansion via ``_full_hist_key`` toggle;
          all events across all sessions; ``LLM_CAUSE_ENRICHED`` kept
          only on first occurrence per claim.
        - All expand/collapse toggles use ``on_click`` callbacks so the
          dialog stays open without triggering a full ``st.rerun()``.

    Dialog persistence:
        - Do NOT call this function directly from a button handler.
        - Set ``st.session_state["_open_journey_dialog"] = {...}`` and
          let ``app.py`` call this on the next rerun (see module docstring).
        - The Close button is the ONLY place that pops
          ``"_open_journey_dialog"`` from session state.

    Args:
        claim_id (str): Stable identifier for the claim being traced
            (e.g. "CLM-001" or a row-index fallback like "claim_3").
        curr_claim (dict): The raw claim dict from the sheet cache,
            mapping field names to cell-info dicts with at least
            "value", and optionally "modified", "excel_row", "excel_col",
            "header_score", "value_score", "confidence", "from_title".
        selected_sheet (str): Active sheet/tab name. Used to scope
            session-state keys, audit log filtering, and field history lookups.
        active_schema (str | None): Name of the currently active schema
            (e.g. "Standard", "Guidewire"), or None if no schema is active.
            When set, fields are shown in schema-mapped order and the
            pipeline trace includes a SCHEMA MAPPED step.
        _llm_map_result (dict | None): LLM column-mapping result dict
            as produced by the LLM mapping call. Expected keys:
            - "mappings"    (dict)  : {source_col: standard_field}
            - "_reasoning"  (dict)  : {source_col: reasoning_string}
            - "_timestamp"  (str)   : ISO timestamp of the LLM call.
            - "_model"      (str)   : model name used.
            Pass None or an empty dict when no LLM mapping was performed.

    Returns:
        None. All output is written directly to the Streamlit dialog.

    Side effects:
        - Initialises session-state keys:
          ``_audit_expand_key`` (set of expanded card IDs),
          ``_full_hist_key`` (bool for full-history toggle).
        - Toggles the above keys via ``on_click`` callbacks on expand
          and history buttons — no ``st.rerun()`` from within toggles.
        - Pops ``st.session_state["_open_journey_dialog"]`` and calls
          ``st.rerun()`` when the Close button is clicked.

    Dependencies:
        - ``modules.audit._load_audit_log``               : full audit event list.
        - ``modules.field_history._get_field_history``    : per-field edit history.
        - ``modules.schema_mapping.map_claim_to_schema``  : schema field mapping.
        - ``config.schemas.SCHEMAS``                      : schema definitions.
        - ``streamlit``                                   : dialog and widgets.
    """
    import json as _json
    import datetime as _dt
    from modules.audit import _load_audit_log
    from modules.field_history import _get_field_history
    from modules.schema_mapping import map_claim_to_schema

    # ── Session-state keys ────────────────────────────────────────────────────
    _audit_expand_key = f"_audit_expanded_{selected_sheet}_{claim_id}"
    _full_hist_key    = f"_audit_fullhist_{selected_sheet}_{claim_id}"
    if _audit_expand_key not in st.session_state:
        st.session_state[_audit_expand_key] = set()
    if _full_hist_key not in st.session_state:
        st.session_state[_full_hist_key] = False

    # ── Timestamps ────────────────────────────────────────────────────────────
    _ts_dialog_open = _dt.datetime.now()
    _ts_fmt = lambda d: d.strftime("%H:%M:%S.%f")[:-3]

    st.markdown(
        f"<div style='font-size:18px;font-weight:700;color:#e8e7ff;margin-bottom:4px;'>"
        f"🔍 Transformation Journey</div>"
        f"<div style='font-size:12px;color:#a0a0c8;font-family:monospace;margin-bottom:4px;'>"
        f"Claim {claim_id} · Sheet: {selected_sheet}"
        + (f" · Schema: {active_schema}" if active_schema else "")
        + f"</div>"
        f"<div style='font-size:10px;color:#555;font-family:monospace;margin-bottom:16px;'>"
        f"⏱ Dialog opened at {_ts_fmt(_ts_dialog_open)}</div>",
        unsafe_allow_html=True,
    )

    # ── Load audit events ─────────────────────────────────────────────────────
    _all_audit   = _load_audit_log()
    _claim_audit = [
        e for e in _all_audit
        if e.get("claim_id") == claim_id and e.get("sheet") == selected_sheet
    ]

    # ── Unpack LLM result ─────────────────────────────────────────────────────
    _ts_llm_unpack   = _dt.datetime.now()
    _llm_mappings    = (_llm_map_result or {}).get("mappings", {})
    _llm_reasoning   = (_llm_map_result or {}).get("_reasoning", {})
    _llm_called_at   = (_llm_map_result or {}).get("_timestamp", None)
    _llm_model       = (_llm_map_result or {}).get("_model", "see .env")
    _llm_reverse     = {v: k for k, v in _llm_mappings.items()}
    _llm_source_cols = set(_llm_mappings.keys())

    # ── Schema mapping ────────────────────────────────────────────────────────
    _ts_schema_map = _dt.datetime.now()
    _mapped: dict = {}
    if active_schema:
        from config.schemas import SCHEMAS
        if active_schema in SCHEMAS:
            _mapped = map_claim_to_schema(curr_claim, active_schema, {}, _llm_map_result)
    _ts_schema_map_done = _dt.datetime.now()
    _schema_map_ms = int((_ts_schema_map_done - _ts_schema_map).total_seconds() * 1000)

    # ── Pipeline trace banner ─────────────────────────────────────────────────
    _pipeline_steps = []
    _pipeline_steps.append(
        f"<div style='display:flex;align-items:center;gap:6px;padding:5px 0;"
        f"border-bottom:1px solid #1e1e32;'>"
        f"<span style='min-width:140px;font-size:10px;color:#34d399;font-weight:700;font-family:monospace;'>"
        f"📂 FILE PARSED</span>"
        f"<span style='font-size:10px;color:#555;font-family:monospace;'>→</span>"
        f"<span style='font-size:10px;color:#a0a0c8;font-family:monospace;'>"
        f"Claims read from the uploaded spreadsheet"
        f"<span style='margin-left:auto;font-size:10px;color:#555;font-family:monospace;'>"
        f" {_ts_fmt(_ts_dialog_open)}</span></div>"
    )
    if active_schema:
        _pipeline_steps.append(
            f"<div style='display:flex;align-items:center;gap:6px;padding:5px 0;"
            f"border-bottom:1px solid #1e1e32;'>"
            f"<span style='min-width:140px;font-size:10px;color:#4f9cf9;font-weight:700;font-family:monospace;'>"
            f"🗂 SCHEMA MAPPED</span>"
            f"<span style='font-size:10px;color:#555;font-family:monospace;'>→</span>"
            f"<span style='font-size:10px;color:#a0a0c8;font-family:monospace;'>"
            f"Fields matched to the {active_schema} schema template"
            f"<span style='margin-left:auto;font-size:10px;color:#555;font-family:monospace;'>"
            f" {_ts_fmt(_ts_schema_map)} ({_schema_map_ms}ms)</span></div>"
        )
    if _llm_mappings:
        _llm_ts_display = _llm_called_at if _llm_called_at else _ts_fmt(_ts_llm_unpack)
        _pipeline_steps.append(
            f"<div style='display:flex;align-items:center;gap:6px;padding:5px 0;"
            f"border-bottom:1px solid #1e1e32;'>"
            f"<span style='min-width:140px;font-size:10px;color:#f5c842;font-weight:700;font-family:monospace;'>"
            f"🤖 LLM CALLED</span>"
            f"<span style='font-size:10px;color:#555;font-family:monospace;'>→</span>"
            f"<span style='font-size:10px;color:#a0a0c8;font-family:monospace;'>"
            f"AI resolved {len(_llm_mappings)} unrecognised column(s) to known fields"
            f"{len(_llm_mappings)} column(s) resolved</span>"
            f"<span style='margin-left:auto;font-size:10px;color:#555;font-family:monospace;'>"
            f" {_llm_ts_display}</span></div>"
        )
    _edit_count = sum(1 for e in _claim_audit if e.get("event") == "FIELD_EDITED")
    if _edit_count:
        _last_edit_ts = max(
            (e.get("timestamp", "")[:19] for e in _claim_audit if e.get("event") == "FIELD_EDITED"),
            default=None,
        )
        _pipeline_steps.append(
            f"<div style='display:flex;align-items:center;gap:6px;padding:5px 0;'>"
            f"<span style='min-width:140px;font-size:10px;color:#f5c842;font-weight:700;font-family:monospace;'>"
            f"✏ USER EDITS</span>"
            f"<span style='font-size:10px;color:#555;font-family:monospace;'>→</span>"
            f"<span style='font-size:10px;color:#a0a0c8;font-family:monospace;'>"
            f"{_edit_count} field(s) manually updated by the user"
            f"<span style='margin-left:auto;font-size:10px;color:#555;font-family:monospace;'>"
            f" {_last_edit_ts.replace('T',' ') if _last_edit_ts else '—'}</span></div>"
        )

    st.markdown(
        f"<div style='background:#0d0d1a;border:1px solid #2a2a45;border-radius:8px;"
        f"padding:10px 14px;margin-bottom:16px;'>"
        f"<div style='font-size:10px;font-weight:700;color:#a0a0c8;font-family:monospace;"
        f"text-transform:uppercase;letter-spacing:1.5px;margin-bottom:8px;'>⚡ Pipeline Trace</div>"
        + "".join(_pipeline_steps)
        + "</div>",
        unsafe_allow_html=True,
    )

    # ── Field timeline ────────────────────────────────────────────────────────
    st.markdown(
        "<div style='display:flex;align-items:center;gap:8px;margin-bottom:12px;'>"
        "<div style='font-size:11px;font-weight:700;color:#a0a0c8;font-family:monospace;"
        "text-transform:uppercase;letter-spacing:1.5px;'>Field Transformation Timeline</div>"
        "<div style='flex:1;height:1px;background:linear-gradient(90deg,#2a2a45,transparent);'></div>"
        "</div>",
        unsafe_allow_html=True,
    )

    fields_to_show = list(_mapped.keys()) if _mapped else list(curr_claim.keys())

    for field in fields_to_show:
        _ts_field = _dt.datetime.now()

        if _mapped and field in _mapped:
            m          = _mapped[field]
            raw_val    = m["info"].get("value", "")
            excel_col  = m.get("excel_field", field)
            hdr_score  = m.get("header_score", 0)
            val_score  = m.get("value_score", 0)
            conf       = m.get("confidence", 0)
            from_title = m.get("from_title", False)
            llm_mapped = bool(m.get("llm_mapped", False)) or (field in _llm_reverse)
            if llm_mapped and field in _llm_reverse:
                excel_col = _llm_reverse[field]
        else:
            if field not in curr_claim:
                continue
            info       = curr_claim[field]
            raw_val    = info.get("value", "")
            excel_col  = field
            hdr_score  = info.get("header_score", 0)
            val_score  = info.get("value_score", 0)
            conf       = info.get("confidence", 0)
            from_title = info.get("from_title", False)
            llm_mapped = (field in _llm_source_cols) or (field in _llm_reverse)

        mk_schema = f"mod_{selected_sheet}_{claim_id}_schema_{field}"
        mk_plain  = f"mod_{selected_sheet}_{claim_id}_{field}"
        cur_val   = (
            st.session_state.get(mk_schema)
            or st.session_state.get(mk_plain)
            or raw_val
        )
        edits     = _get_field_history(selected_sheet, claim_id, field)
        is_edited = cur_val != raw_val

        if from_title:
            method       = "TITLE ROW";    method_color = "#a78bfa"; method_icon = "📋"
            method_fn    = "parsing.py · extract_title_fields()"
        elif llm_mapped:
            method       = "LLM MAPPED";   method_color = "#f5c842"; method_icon = "🤖"
            method_fn    = f"modules.llm · llm_map_unknown_fields() → {_llm_model}"
        elif hdr_score >= 90:
            method       = "EXACT MATCH";  method_color = "#34d399"; method_icon = "✓"
            method_fn    = "modules.schema_mapping · _header_match_score()"
        elif hdr_score >= 65:
            method       = "FUZZY MATCH";  method_color = "#4f9cf9"; method_icon = "~"
            method_fn    = "modules.schema_mapping · _header_match_score() [fuzzy]"
        elif hdr_score > 0:
            method       = "PARTIAL MATCH"; method_color = "#94a3b8"; method_icon = "≈"
            method_fn    = "modules.schema_mapping · _header_match_score() [partial]"
        else:
            method       = "DIRECT";       method_color = "#a0a0c8"; method_icon = "→"
            method_fn    = "modules.parsing · direct column read"

        conf_color   = "#34d399" if conf >= 80 else "#f5c842" if conf >= 50 else "#f87171"
        _empty_html  = "<span style='color:#555;'>(empty)</span>"
        _display_val = raw_val if raw_val else _empty_html
        _field_ts    = _ts_fmt(_ts_field)

        steps_html = ""

        # Step 1 — Extraction
        steps_html += (
            f"<div style='display:flex;align-items:flex-start;gap:10px;margin-bottom:8px;'>"
            f"<div style='min-width:22px;height:22px;border-radius:50%;background:#1e3a2a;"
            f"border:2px solid #34d399;display:flex;align-items:center;justify-content:center;"
            f"font-size:10px;color:#34d399;font-weight:bold;flex-shrink:0;'>1</div>"
            f"<div style='flex:1;'>"
            f"<div style='display:flex;align-items:center;gap:8px;'>"
            f"<div style='font-size:11px;font-weight:700;color:#34d399;font-family:monospace;"
            f"text-transform:uppercase;letter-spacing:1px;'>Extracted from Document</div>"
            f"<span style='font-size:9px;color:#555;font-family:monospace;margin-left:auto;'>"
            f"⏱ {_field_ts} · modules.parsing</span></div>"
            f"<div style='font-size:12px;color:#a0a0c8;margin-top:2px;'>"
            f"Column: <code style='color:#e8e7ff;background:#1a1a2e;padding:1px 5px;"
            f"border-radius:3px;'>{excel_col}</code></div>"
            f"<div style='font-size:13px;color:#e8e7ff;font-family:monospace;"
            f"background:#12121c;border:1px solid #2a2a45;border-radius:4px;"
            f"padding:4px 8px;margin-top:4px;word-break:break-all;'>{_display_val}</div>"
            f"</div></div>"
        )
        steps_html += "<div style='margin-left:11px;border-left:2px dashed #2a2a45;height:8px;margin-bottom:8px;'></div>"

        # Step 2 — Mapping
        steps_html += (
            f"<div style='display:flex;align-items:flex-start;gap:10px;margin-bottom:8px;'>"
            f"<div style='min-width:22px;height:22px;border-radius:50%;background:#1a2540;"
            f"border:2px solid {method_color};display:flex;align-items:center;justify-content:center;"
            f"font-size:10px;color:{method_color};font-weight:bold;flex-shrink:0;'>2</div>"
            f"<div style='flex:1;'>"
            f"<div style='display:flex;align-items:center;gap:8px;'>"
            f"<div style='font-size:11px;font-weight:700;color:{method_color};font-family:monospace;"
            f"text-transform:uppercase;letter-spacing:1px;'>{method_icon} {method}</div>"
            f"<span style='font-size:9px;color:#555;font-family:monospace;margin-left:auto;'>"
            f"⏱ {_field_ts} · {method_fn}</span></div>"
        )
        if llm_mapped:
            _src_col = _llm_reverse.get(field, field)
            _reason  = _llm_reasoning.get(_src_col, "")
            _llm_ts_step = _llm_called_at if _llm_called_at else "—"
            steps_html += (
                f"<div style='font-size:12px;color:#a0a0c8;margin-top:2px;'>"
                f"Source column: <code style='color:#f5c842;background:#1a1a2e;"
                f"padding:1px 5px;border-radius:3px;'>{_src_col}</code>"
            )
            if active_schema and _src_col != field:
                steps_html += (
                    f" → mapped to <code style='color:#f5c842;background:#1a1a2e;"
                    f"padding:1px 5px;border-radius:3px;'>{field}</code>"
                )
            steps_html += (
                f"</div>"
                f"<div style='font-size:10px;color:#555;font-family:monospace;margin-top:2px;'>"
                f"LLM called at: {_llm_ts_step} · model: {_llm_model}</div>"
            )
            if _reason:
                steps_html += (
                    f"<div style='font-size:11px;color:#a0a0c8;font-style:italic;"
                    f"margin-top:3px;padding:4px 8px;background:#1a1a2e;"
                    f"border-left:2px solid #f5c842;border-radius:0 4px 4px 0;'>"
                    f"LLM reasoning: {_reason}</div>"
                )
        elif hdr_score > 0:
            steps_html += (
                f"<div style='font-size:12px;color:#a0a0c8;margin-top:2px;'>"
                f"Header similarity: <span style='color:{method_color};font-weight:700;'>"
                f"{hdr_score}%</span> · "
                f"Value quality: <span style='color:{conf_color};font-weight:700;'>"
                f"{val_score}%</span> · "
                f"Overall confidence: <span style='color:{conf_color};font-weight:700;'>"
                f"{conf}%</span></div>"
            )
        else:
            steps_html += (
                f"<div style='font-size:12px;color:#a0a0c8;margin-top:2px;'>"
                f"Column name matches field directly (no fuzzy scoring applied)</div>"
            )
        steps_html += "</div></div>"

        # Edit steps
        for i, edit in enumerate(edits):
            steps_html += (
                "<div style='margin-left:11px;border-left:2px dashed #2a2a45;height:8px;margin-bottom:8px;'></div>"
                f"<div style='display:flex;align-items:flex-start;gap:10px;margin-bottom:8px;'>"
                f"<div style='min-width:22px;height:22px;border-radius:50%;background:#2a1a10;"
                f"border:2px solid #f5c842;display:flex;align-items:center;justify-content:center;"
                f"font-size:10px;color:#f5c842;font-weight:bold;flex-shrink:0;'>{i+3}</div>"
                f"<div style='flex:1;'>"
                f"<div style='display:flex;align-items:center;gap:8px;'>"
                f"<div style='font-size:11px;font-weight:700;color:#f5c842;font-family:monospace;"
                f"text-transform:uppercase;letter-spacing:1px;'>✏ User Edit</div>"
                f"<span style='font-size:9px;color:#555;font-family:monospace;margin-left:auto;'>"
                f"⏱ {edit['ts']} · ui.claim_panel · _plain_edit_col()</span></div>"
                f"<div style='display:flex;gap:8px;margin-top:4px;align-items:center;'>"
                f"<div style='font-size:12px;color:#f87171;font-family:monospace;"
                f"background:#2a1218;border:1px solid #f87171;border-radius:4px;"
                f"padding:3px 8px;word-break:break-all;flex:1;'>"
                f"<span style='font-size:10px;color:#f87171;'>FROM: </span>{edit['from']}</div>"
                f"<div style='color:#a0a0c8;font-size:14px;'>→</div>"
                f"<div style='font-size:12px;color:#34d399;font-family:monospace;"
                f"background:#0a2a1a;border:1px solid #34d399;border-radius:4px;"
                f"padding:3px 8px;word-break:break-all;flex:1;'>"
                f"<span style='font-size:10px;color:#34d399;'>TO: </span>{edit['to']}</div>"
                f"</div></div></div>"
            )

        if is_edited:
            steps_html += (
                "<div style='margin-left:11px;border-left:2px dashed #2a2a45;height:8px;margin-bottom:8px;'></div>"
                f"<div style='display:flex;align-items:flex-start;gap:10px;margin-bottom:4px;'>"
                f"<div style='min-width:22px;height:22px;border-radius:50%;background:#0a2a1a;"
                f"border:2px solid #34d399;display:flex;align-items:center;justify-content:center;"
                f"font-size:10px;color:#34d399;flex-shrink:0;'>✓</div>"
                f"<div style='flex:1;'>"
                f"<div style='display:flex;align-items:center;gap:8px;'>"
                f"<div style='font-size:11px;font-weight:700;color:#34d399;font-family:monospace;"
                f"text-transform:uppercase;letter-spacing:1px;'>Final Value</div>"
                f"<span style='font-size:9px;color:#555;font-family:monospace;margin-left:auto;'>"
                f"⏱ {_ts_fmt(_dt.datetime.now())} · export ready</span></div>"
                f"<div style='font-size:13px;color:#34d399;font-family:monospace;"
                f"background:#0a2a1a;border:1px solid #34d399;border-radius:4px;"
                f"padding:4px 8px;margin-top:4px;word-break:break-all;'>{cur_val}</div>"
                f"</div></div>"
            )

        border_color = "#f5c842" if is_edited else "#2a2a45"
        st.markdown(
            f"<div style='background:#17172a;border:1px solid {border_color};"
            f"border-radius:8px;padding:12px 14px;margin-bottom:10px;'>"
            f"<div style='font-size:12px;font-weight:700;color:#e8e7ff;font-family:monospace;"
            f"text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;"
            f"display:flex;align-items:center;gap:8px;'>"
            f"{field}"
            + (
                f"<span style='font-size:9px;background:#2a1a10;color:#f5c842;"
                f"padding:1px 6px;border-radius:3px;border:1px solid #f5c842;'>"
                f"MODIFIED</span>" if is_edited else ""
            )
            + f"</div>{steps_html}</div>",
            unsafe_allow_html=True,
        )

    # ══════════════════════════════════════════════════════════════════════════
    # AUDIT LOG SECTION
    # ══════════════════════════════════════════════════════════════════════════
    if _claim_audit:
        st.markdown("---")

        _ev_cfg = {
            "FIELD_EDITED":             ("#4f9cf9", "✏",  "Field edited"),
            "FIELD_ADDED":              ("#a78bfa", "＋", "Custom field added"),
            "EXPORT_GENERATED":         ("#34d399", "⬇",  "Export generated"),
            "FILE_UPLOADED":            ("#f5c842", "📂", "File uploaded"),
            "SCHEMA_CHANGED":           ("#94a3b8", "🗂",  "Schema changed"),
            "LLM_CAUSE_ENRICHED":       ("#6b7280", "🤖", "LLM enriched"),
            "CLAIM_DUPLICATE_DETECTED": ("#f87171", "⚠",  "Duplicate detected"),
        }

        _USER_EVENTS = {"FIELD_EDITED", "FIELD_ADDED", "EXPORT_GENERATED"}
        _session_start = st.session_state.get("_session_start", "")

        _seen_llm = False
        _deduped_audit: list = []
        for _e in _claim_audit:
            if _e.get("event") == "LLM_CAUSE_ENRICHED":
                if not _seen_llm:
                    _deduped_audit.append(_e)
                    _seen_llm = True
            else:
                _deduped_audit.append(_e)

        _session_user_events = [
            e for e in _deduped_audit
            if e.get("event") in _USER_EVENTS
            and e.get("timestamp", "") >= _session_start
        ]

        _full_events = _deduped_audit
        _show_full = st.session_state[_full_hist_key]

        _type_counts: dict[str, int] = {}
        for _e in _session_user_events:
            _t = _e.get("event", "EVENT")
            _type_counts[_t] = _type_counts.get(_t, 0) + 1

        _summary_pills = "".join(
            f"<span style='background:{_ev_cfg.get(_t,('#a0a0c8','•',''))[0]}18;"
            f"border:1px solid {_ev_cfg.get(_t,('#a0a0c8','•',''))[0]}55;"
            f"border-radius:20px;padding:2px 8px;font-size:10px;"
            f"color:{_ev_cfg.get(_t,('#a0a0c8','•',_t))[0]};font-family:monospace;margin-right:4px;'>"
            f"{_ev_cfg.get(_t,('#a0a0c8','•',_t))[1]} {_n} {_t.replace('_',' ').lower()}</span>"
            for _t, _n in _type_counts.items()
        )

        _hdr_col, _btn_col = st.columns([7, 3])
        with _hdr_col:
            st.markdown(
                f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:6px;'>"
                f"<div style='font-size:11px;font-weight:700;color:#a0a0c8;font-family:monospace;"
                f"text-transform:uppercase;letter-spacing:1.5px;'>📋 Audit Log"
                f"<span style='font-size:9px;color:#555;margin-left:6px;font-weight:400;'>"
                f"(this session)</span></div>"
                f"<div style='flex:1;'>{_summary_pills}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )
        with _btn_col:
            _full_label = "▲ Hide Full History" if _show_full else "▼ View Full History"

            def _toggle_full_hist(_key=_full_hist_key):
                st.session_state[_key] = not st.session_state[_key]

            st.button(
                _full_label,
                key=f"toggle_full_hist_{claim_id}",
                use_container_width=True,
                on_click=_toggle_full_hist,
            )

        def _render_audit_rows(events: list, id_prefix: str) -> None:
            """
            Render a list of audit events as expandable row cards.

            Each event is displayed as a compact single-line card with a
            colour-coded left border, event type, timestamp, and a brief
            inline detail string. A ▼/▲ toggle button to the right of
            each card expands or collapses a detail panel showing every
            key-value pair in the event dict (excluding null/empty values).

            Toggle state is stored in
            ``st.session_state[_audit_expand_key]`` as a set of expanded
            card IDs. All toggles use ``on_click`` callbacks so the
            dialog stays open without triggering a ``st.rerun()``.

            Args:
                events (list[dict]): Ordered list of audit event dicts to
                    render. Each dict must contain at least an "event"
                    (str) key. Optional keys rendered in detail panels
                    include "timestamp", "field", "original", "new_value",
                    "export_type", "records", "cause_of_loss".
                id_prefix (str): A unique string prefix used to namespace
                    card toggle keys (e.g. "user_Q1_CLM-001" or
                    "full_Q1_CLM-001"). Prevents key collisions between
                    the session view and the full history view when both
                    are visible simultaneously.

            Returns:
                None. Writes directly to the Streamlit dialog.

            Side effects:
                - Reads and mutates
                  ``st.session_state[_audit_expand_key]`` (the set of
                  expanded card IDs) via ``on_click`` callbacks.
                - Does NOT call ``st.rerun()`` — all state changes happen
                  through the on_click mechanism.
            """
            for _ei, _event in enumerate(events):
                _ev_type  = _event.get("event", "EVENT")
                _ev_ts    = _event.get("timestamp", "")[:19].replace("T", " ")
                _ev_field = _event.get("field", "")
                _ev_from  = _event.get("original", "")
                _ev_to    = _event.get("new_value", "")
                _ev_recs  = _event.get("records", "")
                _ev_etype = _event.get("export_type", "")

                _cfg      = _ev_cfg.get(_ev_type, ("#a0a0c8", "•", _ev_type))
                _ev_color = _cfg[0]
                _ev_icon  = _cfg[1]

                _detail = ""
                if _ev_type == "FIELD_EDITED" and _ev_field:
                    _sf   = str(_ev_from)[:22] + ("…" if len(str(_ev_from)) > 22 else "")
                    _st_v = str(_ev_to)[:22]   + ("…" if len(str(_ev_to))   > 22 else "")
                    _detail = (
                        f"<span style='color:#e8e7ff;'>{_ev_field}</span> "
                        f"<span style='color:#f87171;'>{_sf}</span>"
                        f"<span style='color:#555;'> → </span>"
                        f"<span style='color:#34d399;'>{_st_v}</span>"
                    )
                elif _ev_type == "FIELD_ADDED" and _ev_field:
                    _detail = f"<span style='color:#a78bfa;'>{_ev_field}</span>"
                elif _ev_type == "EXPORT_GENERATED":
                    _detail = (
                        f"<span style='color:#34d399;'>{_ev_etype}</span>"
                        + (f"<span style='color:#555;'> · {_ev_recs} records</span>" if _ev_recs else "")
                    )
                elif _ev_type == "LLM_CAUSE_ENRICHED":
                    _cause = _event.get("cause_of_loss", "")
                    _detail = (
                        f"<span style='color:#6b7280;'>cause: </span>"
                        f"<span style='color:#a0a0c8;'>{_cause}</span>"
                        if _cause else
                        "<span style='color:#6b7280;font-style:italic;'>first enrichment only</span>"
                    )
                elif _ev_type == "CLAIM_DUPLICATE_DETECTED":
                    _detail = "<span style='color:#f87171;font-style:italic;'>duplicate flag raised</span>"

                _card_key = f"{id_prefix}_{_ei}"
                _expanded = _card_key in st.session_state[_audit_expand_key]
                _btn_lbl  = "▲" if _expanded else "▼"

                def _toggle_card(_ck=_card_key, _ek=_audit_expand_key):
                    if _ck in st.session_state[_ek]:
                        st.session_state[_ek].discard(_ck)
                    else:
                        st.session_state[_ek].add(_ck)

                _row_col, _xbtn_col = st.columns([10, 1])
                with _row_col:
                    st.markdown(
                        f"<div style='background:#12121c;border:1px solid #2a2a45;"
                        f"border-left:3px solid {_ev_color};border-radius:6px;"
                        f"padding:8px 12px;font-family:monospace;font-size:11px;"
                        f"display:flex;align-items:center;gap:8px;'>"
                        f"<span style='color:{_ev_color};font-weight:700;min-width:16px;'>{_ev_icon}</span>"
                        f"<span style='color:{_ev_color};font-weight:700;min-width:175px;"
                        f"white-space:nowrap;'>{_ev_type}</span>"
                        f"<span style='color:#555;'>·</span>"
                        f"<span style='color:#6b7280;min-width:135px;white-space:nowrap;'>{_ev_ts}</span>"
                        f"<span style='color:#555;'>·</span>"
                        f"<span style='flex:1;'>{_detail}</span>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                with _xbtn_col:
                    st.button(
                        _btn_lbl,
                        key=f"audit_expand_{_card_key}",
                        use_container_width=True,
                        help="Expand / collapse full details",
                        on_click=_toggle_card,
                    )

                if _expanded:
                    _detail_rows = ""
                    for _k, _v in _event.items():
                        if _k == "event" or _v in (None, "", []):
                            continue
                        _vs = str(_v)
                        if _k == "original":                  _vc = "#f87171"
                        elif _k == "new_value":               _vc = "#34d399"
                        elif _k == "timestamp":               _vc = "#6b7280"
                        elif _k in ("field", "export_type"):  _vc = "#e8e7ff"
                        else:                                 _vc = "#a0a0c8"
                        _detail_rows += (
                            f"<div style='display:flex;gap:12px;padding:4px 0;"
                            f"border-bottom:1px solid #1a1a2e;'>"
                            f"<span style='min-width:120px;font-size:10px;color:#555;"
                            f"font-family:monospace;text-transform:uppercase;'>{_k}</span>"
                            f"<span style='font-size:11px;color:{_vc};font-family:monospace;"
                            f"word-break:break-all;'>{_vs}</span>"
                            f"</div>"
                        )
                    st.markdown(
                        f"<div style='background:#0d0d1a;border:1px solid {_ev_color}44;"
                        f"border-left:3px solid {_ev_color};border-radius:0 0 6px 6px;"
                        f"padding:10px 14px;margin-top:-6px;margin-bottom:8px;'>"
                        f"{_detail_rows}</div>",
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown("<div style='margin-bottom:6px;'></div>", unsafe_allow_html=True)

        if not _session_user_events:
            st.markdown(
                "<div style='color:#555;font-size:12px;font-family:monospace;"
                "padding:10px 0;font-style:italic;'>"
                "No user actions recorded in this session yet for this claim.</div>",
                unsafe_allow_html=True,
            )
        else:
            _render_audit_rows(_session_user_events, f"user_{selected_sheet}_{claim_id}")

        if _show_full:
            _llm_count_raw = sum(1 for e in _claim_audit if e.get("event") == "LLM_CAUSE_ENRICHED")
            _suppressed    = _llm_count_raw - (1 if _llm_count_raw > 0 else 0)
            _note = (
                f" · {_suppressed} duplicate LLM_CAUSE_ENRICHED event(s) suppressed"
                if _suppressed > 0 else ""
            )
            st.markdown(
                f"<div style='background:#0d0d1a;border:1px solid #2a2a45;border-radius:8px;"
                f"padding:10px 14px;margin-top:8px;'>"
                f"<div style='font-size:10px;font-weight:700;color:#6b7280;font-family:monospace;"
                f"text-transform:uppercase;letter-spacing:1.5px;margin-bottom:10px;'>"
                f"🕓 Full History — {len(_full_events)} event(s){_note}</div>",
                unsafe_allow_html=True,
            )
            _render_audit_rows(_full_events, f"full_{selected_sheet}_{claim_id}")
            st.markdown("</div>", unsafe_allow_html=True)

    if st.button("Close", type="primary", use_container_width=True):
        st.session_state.pop("_open_journey_dialog", None)
        st.rerun()
